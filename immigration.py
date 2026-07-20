from flask import jsonify, redirect, render_template, request, send_file, session, url_for
import csv
import io
import json
import os
import re
import tempfile
from datetime import datetime, timedelta

from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from werkzeug.utils import secure_filename

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

try:
    import pytesseract
except Exception:
    pytesseract = None


IMMIGRATION_WORKSPACE = 'immigration'
BASE_DIR = os.path.dirname(__file__)
STATIC_ROOT = os.path.join(BASE_DIR, 'static')
DEFAULT_UPLOAD_ROOT = os.path.join(STATIC_ROOT, 'uploads', 'immigration')
UPLOAD_ROOT = os.getenv('IMMIGRATION_UPLOAD_ROOT') or (
    DEFAULT_UPLOAD_ROOT if os.access(STATIC_ROOT, os.W_OK) else os.path.join(tempfile.gettempdir(), 'bdj-ledger', 'immigration')
)
ALLOWED_IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.webp'}
ALLOWED_DOCUMENT_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.webp', '.pdf', '.doc', '.docx', '.xls', '.xlsx'}
APPLICANT_TITLES = ['Mr.', 'Mrs.', 'Miss', 'Ms.', 'Master', 'Doctor', 'Prof']
TITLE_NORMALIZATION = {
    'mr': 'Mr.',
    'mr.': 'Mr.',
    'mrs': 'Mrs.',
    'mrs.': 'Mrs.',
    'miss': 'Miss',
    'ms': 'Ms.',
    'ms.': 'Ms.',
    'master': 'Master',
    'dr': 'Doctor',
    'dr.': 'Doctor',
    'doctor': 'Doctor',
    'prof': 'Prof',
    'prof.': 'Prof',
    'professor': 'Prof',
}
PASSPORT_TYPES = ['Standard', 'Service', 'Official']
GENDER_OPTIONS = [('MALE', 'Male'), ('FEMALE', 'Female')]
MARITAL_STATUS_OPTIONS = [('SINGLE', 'Single'), ('MARRIED', 'Married'), ('WIDOWED', 'Widowed'), ('DIVORCED', 'Divorced')]
VISA_CATEGORIES = [
    'F3B - Transit Visa',
    'F4A - Business',
    'F4B - Business Visa (Multiple Entry)',
    'F5A - Tourism Visa',
    'F6A - Visiting Visa (Single Entry)',
    'F7E - Sports Visa',
    'F7F - Creative Arts Visa',
    'F7G - Study Tour Visa',
    'F7H - Academic Exchange Programme Visa',
    'F71 - International Cultural Exchange Visa',
    'F7K - Emergency/Relief Work Visa',
    'F9A - Returning holders of foreign Passports who are Nigerians by Birth',
    'F9B - Returning Holders of Foreign Passports (Nigerian by Birth)',
]
REASONS = ['Business Meeting', 'Business Talks', 'Conference', 'Visit']
JOURNEY_PURPOSES = [
    'Business Meeting',
    'Business Talks',
    'Conference',
    'Tourism',
    'Visit',
    'Training',
    'Transit',
    'Academic Exchange',
    'Study Tour',
    'Sports',
    'Creative Arts',
    'Emergency/Relief Work',
    'Returning Nigerian by Birth',
]
ARRIVAL_CHANNELS = ['Air', 'Land', 'Sea']
DEFAULT_ENTRY_PORTS = {
    'Air': [
        'Nnamdi Azikiwe international Airport, Abuja',
        'Mallam Aminu Kano Airport, Kano',
        'Murtala Mohammed Airport, Lagos',
        'Margret Ekpo Airport, Calabar',
        'PortHarcourt International Airport, Rivers',
        'Akanu Ibiam International Airport, Enugu',
    ],
    'Land': ['Seme Border', 'Jibiya Border', 'Idiroko Border', 'Mfum Border'],
    'Sea': ['Apapa Seaport', 'Tin Can Island Port', 'Onne Port', 'Calabar Port'],
}
DOCUMENT_TYPES = [
    'Passport',
    'Photograph',
    'Invitation',
    'Hotel Reservation',
    'Flight Ticket',
    'Bank Statement',
    'CAC Certificate',
    'Host Letter',
    'Additional Documents',
]
APPLICANT_STATUSES = ['Draft', 'Pending', 'In Review', 'Validated', 'Completed']
COUNTRY_OPTIONS = [
    'Afghanistan', 'Albania', 'Algeria', 'Angola', 'Argentina', 'Armenia', 'Australia', 'Austria', 'Azerbaijan',
    'Bahrain', 'Bangladesh', 'Belarus', 'Belgium', 'Benin', 'Bolivia', 'Bosnia and Herzegovina', 'Botswana', 'Brazil',
    'Brunei', 'Bulgaria', 'Burkina Faso', 'Burundi', 'Cambodia', 'Cameroon', 'Canada', 'Cape Verde', 'Central African Republic',
    'Chad', 'Chile', 'China', 'Colombia', 'Comoros', 'Congo', 'Costa Rica', 'Croatia', 'Cuba', 'Cyprus', 'Czech Republic',
    'Denmark', 'Djibouti', 'Dominican Republic', 'DR Congo', 'Ecuador', 'Egypt', 'El Salvador', 'Equatorial Guinea', 'Eritrea',
    'Estonia', 'Eswatini', 'Ethiopia', 'Fiji', 'Finland', 'France', 'Gabon', 'Gambia', 'Georgia', 'Germany', 'Ghana', 'Greece',
    'Guatemala', 'Guinea', 'Guinea-Bissau', 'Guyana', 'Haiti', 'Honduras', 'Hungary', 'Iceland', 'India', 'Indonesia', 'Iran',
    'Iraq', 'Ireland', 'Israel', 'Italy', 'Ivory Coast', 'Jamaica', 'Japan', 'Jordan', 'Kazakhstan', 'Kenya', 'Kuwait',
    'Kyrgyzstan', 'Laos', 'Latvia', 'Lebanon', 'Lesotho', 'Liberia', 'Libya', 'Lithuania', 'Luxembourg', 'Madagascar',
    'Malawi', 'Malaysia', 'Mali', 'Malta', 'Mauritania', 'Mauritius', 'Mexico', 'Moldova', 'Mongolia', 'Montenegro', 'Morocco',
    'Mozambique', 'Myanmar', 'Namibia', 'Nepal', 'Netherlands', 'New Zealand', 'Nicaragua', 'Niger', 'Nigeria', 'North Korea',
    'North Macedonia', 'Norway', 'Oman', 'Pakistan', 'Panama', 'Paraguay', 'Peru', 'Philippines', 'Poland', 'Portugal', 'Qatar',
    'Romania', 'Russia', 'Rwanda', 'Saudi Arabia', 'Senegal', 'Serbia', 'Sierra Leone', 'Singapore', 'Slovakia', 'Slovenia',
    'Somalia', 'South Africa', 'South Korea', 'South Sudan', 'Spain', 'Sri Lanka', 'Sudan', 'Sweden', 'Switzerland', 'Syria',
    'Taiwan', 'Tajikistan', 'Tanzania', 'Thailand', 'Togo', 'Trinidad and Tobago', 'Tunisia', 'Turkey', 'Turkmenistan', 'Uganda',
    'Ukraine', 'United Arab Emirates', 'United Kingdom', 'United States', 'Uruguay', 'Uzbekistan', 'Venezuela', 'Vietnam',
    'Yemen', 'Zambia', 'Zimbabwe'
]
NATIONALITY_OPTIONS = [
    'Afghan', 'Albanian', 'Algerian', 'Angolan', 'Argentine', 'Armenian', 'Australian', 'Austrian', 'Azerbaijani', 'Bahraini',
    'Bangladeshi', 'Belgian', 'Beninese', 'Bolivian', 'Botswanan', 'Brazilian', 'British', 'Bulgarian', 'Burkinabe', 'Burundian',
    'Cambodian', 'Cameroonian', 'Canadian', 'Chadian', 'Chilean', 'Chinese', 'Colombian', 'Congolese', 'Costa Rican', 'Croatian',
    'Cuban', 'Czech', 'Danish', 'Dominican', 'Ecuadorean', 'Egyptian', 'Emirati', 'Eritrean', 'Ethiopian', 'Finnish', 'French',
    'Gabonese', 'Gambian', 'Georgian', 'German', 'Ghanaian', 'Greek', 'Guinean', 'Haitian', 'Hungarian', 'Indian', 'Indonesian',
    'Iranian', 'Iraqi', 'Irish', 'Israeli', 'Italian', 'Ivorian', 'Jamaican', 'Japanese', 'Jordanian', 'Kazakh', 'Kenyan',
    'Kuwaiti', 'Lao', 'Latvian', 'Lebanese', 'Liberian', 'Libyan', 'Lithuanian', 'Malagasy', 'Malawian', 'Malaysian', 'Malian',
    'Maltese', 'Mauritanian', 'Mauritian', 'Mexican', 'Moldovan', 'Mongolian', 'Moroccan', 'Mozambican', 'Namibian', 'Nepalese',
    'Nigerian', 'Nigerien', 'Norwegian', 'Omani', 'Pakistani', 'Panamanian', 'Paraguayan', 'Peruvian', 'Philippine', 'Polish',
    'Portuguese', 'Qatari', 'Romanian', 'Russian', 'Rwandan', 'Saudi', 'Senegalese', 'Serbian', 'Sierra Leonean', 'Singaporean',
    'Slovak', 'Slovenian', 'Somali', 'South African', 'South Korean', 'Spanish', 'Sri Lankan', 'Sudanese', 'Swedish', 'Swiss',
    'Syrian', 'Taiwanese', 'Tajik', 'Tanzanian', 'Thai', 'Togolese', 'Trinidadian', 'Tunisian', 'Turkish', 'Ugandan', 'Ukrainian',
    'American', 'Uruguayan', 'Uzbek', 'Venezuelan', 'Vietnamese', 'Yemeni', 'Zambian', 'Zimbabwean'
]
AIRLINE_PREFIXES = {
    'TK': 'Turkish Airlines',
    'QR': 'Qatar Airways',
    'ET': 'Ethiopian Airlines',
    'BA': 'British Airways',
    'LH': 'Lufthansa',
    'EK': 'Emirates',
    'KL': 'KLM',
    'AF': 'Air France',
    'MS': 'EgyptAir',
    'SN': 'Brussels Airlines',
    'WB': 'RwandAir',
    'KQ': 'Kenya Airways',
    'P4': 'Air Peace',
    'Q9': 'Ibom Air',
    'VM': 'Max Air',
    'W3': 'Arik Air',
    'N2': 'Aero Contractors',
    'AJ': 'Azman Air',
}
TRAVEL_CARRIERS = sorted(set(AIRLINE_PREFIXES.values()))


def ensure_upload_dirs():
    for folder in [
        UPLOAD_ROOT,
        os.path.join(UPLOAD_ROOT, 'companies'),
        os.path.join(UPLOAD_ROOT, 'letterheads'),
        os.path.join(UPLOAD_ROOT, 'signatures'),
        os.path.join(UPLOAD_ROOT, 'documents'),
        os.path.join(UPLOAD_ROOT, 'ocr'),
    ]:
        os.makedirs(folder, exist_ok=True)


def to_storage_key(folder_name, stored_name):
    return f'{folder_name}/{stored_name}'.replace('\\', '/')


def resolve_storage_path(stored_path):
    normalized = (stored_path or '').replace('\\', '/').lstrip('/')
    if not normalized:
        return None
    if os.path.isabs(stored_path):
        return stored_path
    candidates = []
    if normalized.startswith('uploads/immigration/'):
        candidates.append(os.path.join(STATIC_ROOT, normalized))
    if normalized.startswith('static/'):
        candidates.append(os.path.join(BASE_DIR, normalized))
    candidates.append(os.path.join(UPLOAD_ROOT, normalized))
    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate
    return candidates[-1]


def build_file_url(stored_path):
    if not stored_path:
        return ''
    normalized = stored_path.replace('\\', '/').lstrip('/')
    return url_for('immigration_uploaded_file', stored_path=normalized)


def ensure_immigration_schema(cursor, backend):
    if backend == 'postgres':
        statements = [
            '''
            CREATE TABLE IF NOT EXISTS user_preferences (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL,
                pref_key TEXT NOT NULL,
                pref_value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_user_preferences_unique ON user_preferences(user_id, pref_key)',
            '''
            CREATE TABLE IF NOT EXISTS companies (
                id SERIAL PRIMARY KEY,
                model_id INTEGER,
                name TEXT NOT NULL,
                logo_path TEXT,
                email TEXT,
                phone TEXT,
                address TEXT,
                country TEXT,
                default_contact_id INTEGER,
                default_letterhead_id INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_companies_model_name ON companies(model_id, name)',
            '''
            CREATE TABLE IF NOT EXISTS contacts (
                id SERIAL PRIMARY KEY,
                model_id INTEGER,
                company_id INTEGER,
                contact_name TEXT NOT NULL,
                phone TEXT,
                email TEXT,
                address TEXT,
                city TEXT,
                state TEXT,
                postal_code TEXT,
                country TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS letterheads (
                id SERIAL PRIMARY KEY,
                model_id INTEGER,
                company_id INTEGER,
                template_name TEXT NOT NULL,
                template_type TEXT,
                background_image_path TEXT,
                signature_image_path TEXT,
                signatory TEXT,
                designation TEXT,
                scale_percent REAL DEFAULT 100,
                move_left REAL DEFAULT 0,
                move_right REAL DEFAULT 0,
                move_up REAL DEFAULT 0,
                move_down REAL DEFAULT 0,
                rotation REAL DEFAULT 0,
                plain_paper INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS applicants (
                id SERIAL PRIMARY KEY,
                model_id INTEGER,
                title TEXT,
                surname TEXT NOT NULL,
                first_name TEXT NOT NULL,
                middle_name TEXT,
                full_name TEXT,
                passport_type TEXT,
                gender TEXT,
                marital_status TEXT,
                date_of_birth TEXT,
                place_of_birth TEXT,
                passport_number TEXT,
                passport_expiry TEXT,
                nigerian_passport TEXT,
                flight_number TEXT,
                travel_carrier TEXT,
                nationality TEXT,
                visa_type TEXT,
                status TEXT DEFAULT 'Pending',
                country_of_departure TEXT,
                departure_date TEXT,
                arrival_date TEXT,
                arrival_channel TEXT,
                duration_of_stay TEXT,
                port_of_entry TEXT,
                travel_date TEXT,
                email TEXT,
                phone TEXT,
                contact_name TEXT,
                contact_email TEXT,
                contact_phone TEXT,
                contact_address TEXT,
                contact_city TEXT,
                contact_state TEXT,
                contact_country TEXT,
                contact_postal_code TEXT,
                reference_number TEXT,
                company_id INTEGER,
                contact_id INTEGER,
                reason TEXT,
                notes TEXT,
                draft_data TEXT,
                created_by TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            'CREATE INDEX IF NOT EXISTS idx_applicants_model_created ON applicants(model_id, created_at DESC)',
            'CREATE INDEX IF NOT EXISTS idx_applicants_company ON applicants(company_id)',
            'CREATE INDEX IF NOT EXISTS idx_applicants_contact ON applicants(contact_id)',
            '''
            CREATE TABLE IF NOT EXISTS documents (
                id SERIAL PRIMARY KEY,
                model_id INTEGER,
                applicant_id INTEGER NOT NULL,
                document_type TEXT NOT NULL,
                original_name TEXT,
                stored_name TEXT,
                file_path TEXT NOT NULL,
                mime_type TEXT,
                size_bytes BIGINT DEFAULT 0,
                validation_status TEXT DEFAULT 'Pending',
                extracted_data TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS visa_letters (
                id SERIAL PRIMARY KEY,
                model_id INTEGER,
                applicant_id INTEGER,
                company_id INTEGER,
                letterhead_id INTEGER,
                reason TEXT,
                visa_type TEXT,
                nationality TEXT,
                plain_paper INTEGER DEFAULT 0,
                preview_html TEXT,
                pdf_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS automation_sessions (
                id SERIAL PRIMARY KEY,
                model_id INTEGER,
                status TEXT DEFAULT 'Disconnected',
                browser_name TEXT,
                current_applicant_id INTEGER,
                current_page TEXT,
                progress_percent REAL DEFAULT 0,
                details TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS travel_history (
                id SERIAL PRIMARY KEY,
                model_id INTEGER,
                applicant_id INTEGER,
                travel_date TEXT,
                visa_type TEXT,
                destination TEXT,
                status TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS audit_logs (
                id SERIAL PRIMARY KEY,
                model_id INTEGER,
                user_id INTEGER,
                workspace TEXT,
                action TEXT,
                entity_type TEXT,
                entity_id INTEGER,
                details TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS entry_ports (
                id SERIAL PRIMARY KEY,
                channel TEXT NOT NULL,
                port_name TEXT NOT NULL,
                sort_order INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1
            )
            ''',
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_entry_ports_unique ON entry_ports(channel, port_name)',
        ]
    elif backend == 'mysql':
        statements = [
            '''
            CREATE TABLE IF NOT EXISTS user_preferences (
                id INT AUTO_INCREMENT PRIMARY KEY,
                user_id INT NOT NULL,
                pref_key VARCHAR(255) NOT NULL,
                pref_value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY idx_user_preferences_unique (user_id, pref_key)
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS companies (
                id INT AUTO_INCREMENT PRIMARY KEY,
                model_id INT,
                name VARCHAR(255) NOT NULL,
                logo_path TEXT,
                email VARCHAR(255),
                phone VARCHAR(255),
                address TEXT,
                country VARCHAR(255),
                default_contact_id INT,
                default_letterhead_id INT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY idx_companies_model_name (model_id, name)
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS contacts (
                id INT AUTO_INCREMENT PRIMARY KEY,
                model_id INT,
                company_id INT,
                contact_name VARCHAR(255) NOT NULL,
                phone VARCHAR(255),
                email VARCHAR(255),
                address TEXT,
                city VARCHAR(255),
                state VARCHAR(255),
                postal_code VARCHAR(255),
                country VARCHAR(255),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS letterheads (
                id INT AUTO_INCREMENT PRIMARY KEY,
                model_id INT,
                company_id INT,
                template_name VARCHAR(255) NOT NULL,
                template_type VARCHAR(255),
                background_image_path TEXT,
                signature_image_path TEXT,
                signatory VARCHAR(255),
                designation VARCHAR(255),
                scale_percent DOUBLE DEFAULT 100,
                move_left DOUBLE DEFAULT 0,
                move_right DOUBLE DEFAULT 0,
                move_up DOUBLE DEFAULT 0,
                move_down DOUBLE DEFAULT 0,
                rotation DOUBLE DEFAULT 0,
                plain_paper TINYINT(1) DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS applicants (
                id INT AUTO_INCREMENT PRIMARY KEY,
                model_id INT,
                title VARCHAR(64),
                surname VARCHAR(255) NOT NULL,
                first_name VARCHAR(255) NOT NULL,
                middle_name VARCHAR(255),
                full_name TEXT,
                passport_type VARCHAR(255),
                gender VARCHAR(64),
                marital_status VARCHAR(64),
                date_of_birth VARCHAR(64),
                place_of_birth VARCHAR(255),
                passport_number VARCHAR(255),
                passport_expiry VARCHAR(64),
                nigerian_passport VARCHAR(32),
                flight_number VARCHAR(64),
                travel_carrier VARCHAR(255),
                nationality VARCHAR(255),
                visa_type VARCHAR(255),
                status VARCHAR(64) DEFAULT 'Pending',
                country_of_departure VARCHAR(255),
                departure_date VARCHAR(64),
                arrival_date VARCHAR(64),
                arrival_channel VARCHAR(64),
                duration_of_stay VARCHAR(64),
                port_of_entry VARCHAR(255),
                travel_date VARCHAR(64),
                email VARCHAR(255),
                phone VARCHAR(255),
                contact_name VARCHAR(255),
                contact_email VARCHAR(255),
                contact_phone VARCHAR(255),
                contact_address TEXT,
                contact_city VARCHAR(255),
                contact_state VARCHAR(255),
                contact_country VARCHAR(255),
                contact_postal_code VARCHAR(255),
                reference_number VARCHAR(255),
                company_id INT,
                contact_id INT,
                reason VARCHAR(255),
                notes TEXT,
                draft_data LONGTEXT,
                created_by VARCHAR(255),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            'CREATE INDEX idx_applicants_model_created ON applicants(model_id, created_at)',
            '''
            CREATE TABLE IF NOT EXISTS documents (
                id INT AUTO_INCREMENT PRIMARY KEY,
                model_id INT,
                applicant_id INT NOT NULL,
                document_type VARCHAR(255) NOT NULL,
                original_name TEXT,
                stored_name TEXT,
                file_path TEXT NOT NULL,
                mime_type VARCHAR(255),
                size_bytes BIGINT DEFAULT 0,
                validation_status VARCHAR(64) DEFAULT 'Pending',
                extracted_data LONGTEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS visa_letters (
                id INT AUTO_INCREMENT PRIMARY KEY,
                model_id INT,
                applicant_id INT,
                company_id INT,
                letterhead_id INT,
                reason VARCHAR(255),
                visa_type VARCHAR(255),
                nationality VARCHAR(255),
                plain_paper TINYINT(1) DEFAULT 0,
                preview_html LONGTEXT,
                pdf_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS automation_sessions (
                id INT AUTO_INCREMENT PRIMARY KEY,
                model_id INT,
                status VARCHAR(64) DEFAULT 'Disconnected',
                browser_name VARCHAR(255),
                current_applicant_id INT,
                current_page VARCHAR(255),
                progress_percent DOUBLE DEFAULT 0,
                details LONGTEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS travel_history (
                id INT AUTO_INCREMENT PRIMARY KEY,
                model_id INT,
                applicant_id INT,
                travel_date VARCHAR(64),
                visa_type VARCHAR(255),
                destination VARCHAR(255),
                status VARCHAR(64),
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INT AUTO_INCREMENT PRIMARY KEY,
                model_id INT,
                user_id INT,
                workspace VARCHAR(64),
                action VARCHAR(255),
                entity_type VARCHAR(255),
                entity_id INT,
                details LONGTEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS entry_ports (
                id INT AUTO_INCREMENT PRIMARY KEY,
                channel VARCHAR(32) NOT NULL,
                port_name VARCHAR(255) NOT NULL,
                sort_order INT DEFAULT 0,
                is_active TINYINT(1) DEFAULT 1,
                UNIQUE KEY idx_entry_ports_unique (channel, port_name)
            )
            ''',
        ]
    else:
        statements = [
            '''
            CREATE TABLE IF NOT EXISTS user_preferences (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                pref_key TEXT NOT NULL,
                pref_value TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_user_preferences_unique ON user_preferences(user_id, pref_key)',
            '''
            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                name TEXT NOT NULL,
                logo_path TEXT,
                email TEXT,
                phone TEXT,
                address TEXT,
                country TEXT,
                default_contact_id INTEGER,
                default_letterhead_id INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_companies_model_name ON companies(model_id, name)',
            '''
            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                company_id INTEGER,
                contact_name TEXT NOT NULL,
                phone TEXT,
                email TEXT,
                address TEXT,
                city TEXT,
                state TEXT,
                postal_code TEXT,
                country TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS letterheads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                company_id INTEGER,
                template_name TEXT NOT NULL,
                template_type TEXT,
                background_image_path TEXT,
                signature_image_path TEXT,
                signatory TEXT,
                designation TEXT,
                scale_percent REAL DEFAULT 100,
                move_left REAL DEFAULT 0,
                move_right REAL DEFAULT 0,
                move_up REAL DEFAULT 0,
                move_down REAL DEFAULT 0,
                rotation REAL DEFAULT 0,
                plain_paper INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS applicants (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                title TEXT,
                surname TEXT NOT NULL,
                first_name TEXT NOT NULL,
                middle_name TEXT,
                full_name TEXT,
                passport_type TEXT,
                gender TEXT,
                marital_status TEXT,
                date_of_birth TEXT,
                place_of_birth TEXT,
                passport_number TEXT,
                passport_expiry TEXT,
                nigerian_passport TEXT,
                flight_number TEXT,
                travel_carrier TEXT,
                nationality TEXT,
                visa_type TEXT,
                status TEXT DEFAULT 'Pending',
                country_of_departure TEXT,
                departure_date TEXT,
                arrival_date TEXT,
                arrival_channel TEXT,
                duration_of_stay TEXT,
                port_of_entry TEXT,
                travel_date TEXT,
                email TEXT,
                phone TEXT,
                contact_name TEXT,
                contact_email TEXT,
                contact_phone TEXT,
                contact_address TEXT,
                contact_city TEXT,
                contact_state TEXT,
                contact_country TEXT,
                contact_postal_code TEXT,
                reference_number TEXT,
                company_id INTEGER,
                contact_id INTEGER,
                reason TEXT,
                notes TEXT,
                draft_data TEXT,
                created_by TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            'CREATE INDEX IF NOT EXISTS idx_applicants_model_created ON applicants(model_id, created_at DESC)',
            'CREATE INDEX IF NOT EXISTS idx_applicants_company ON applicants(company_id)',
            'CREATE INDEX IF NOT EXISTS idx_applicants_contact ON applicants(contact_id)',
            '''
            CREATE TABLE IF NOT EXISTS documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                applicant_id INTEGER NOT NULL,
                document_type TEXT NOT NULL,
                original_name TEXT,
                stored_name TEXT,
                file_path TEXT NOT NULL,
                mime_type TEXT,
                size_bytes INTEGER DEFAULT 0,
                validation_status TEXT DEFAULT 'Pending',
                extracted_data TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS visa_letters (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                applicant_id INTEGER,
                company_id INTEGER,
                letterhead_id INTEGER,
                reason TEXT,
                visa_type TEXT,
                nationality TEXT,
                plain_paper INTEGER DEFAULT 0,
                preview_html TEXT,
                pdf_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS automation_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                status TEXT DEFAULT 'Disconnected',
                browser_name TEXT,
                current_applicant_id INTEGER,
                current_page TEXT,
                progress_percent REAL DEFAULT 0,
                details TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS travel_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                applicant_id INTEGER,
                travel_date TEXT,
                visa_type TEXT,
                destination TEXT,
                status TEXT,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_id INTEGER,
                user_id INTEGER,
                workspace TEXT,
                action TEXT,
                entity_type TEXT,
                entity_id INTEGER,
                details TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''',
            '''
            CREATE TABLE IF NOT EXISTS entry_ports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                channel TEXT NOT NULL,
                port_name TEXT NOT NULL,
                sort_order INTEGER DEFAULT 0,
                is_active INTEGER DEFAULT 1
            )
            ''',
            'CREATE UNIQUE INDEX IF NOT EXISTS idx_entry_ports_unique ON entry_ports(channel, port_name)',
        ]

    for statement in statements:
        try:
            cursor.execute(statement)
        except Exception:
            if backend == 'mysql' and 'idx_applicants_model_created' in statement:
                pass
            else:
                raise

    for channel, ports in DEFAULT_ENTRY_PORTS.items():
        for index, port_name in enumerate(ports, start=1):
            try:
                if backend == 'postgres':
                    cursor.execute(
                        '''
                        INSERT INTO entry_ports (channel, port_name, sort_order, is_active)
                        VALUES (%s, %s, %s, 1)
                        ON CONFLICT (channel, port_name) DO NOTHING
                        ''',
                        (channel, port_name, index),
                    )
                elif backend == 'mysql':
                    cursor.execute(
                        '''
                        INSERT IGNORE INTO entry_ports (channel, port_name, sort_order, is_active)
                        VALUES (%s, %s, %s, 1)
                        ''',
                        (channel, port_name, index),
                    )
                else:
                    cursor.execute(
                        '''
                        INSERT OR IGNORE INTO entry_ports (channel, port_name, sort_order, is_active)
                        VALUES (?, ?, ?, 1)
                        ''',
                        (channel, port_name, index),
                    )
            except Exception:
                pass

    required_applicant_columns = [
        ('passport_type', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('gender', 'TEXT' if backend != 'mysql' else 'VARCHAR(64)'),
        ('marital_status', 'TEXT' if backend != 'mysql' else 'VARCHAR(64)'),
        ('nigerian_passport', 'TEXT' if backend != 'mysql' else 'VARCHAR(32)'),
        ('flight_number', 'TEXT' if backend != 'mysql' else 'VARCHAR(64)'),
        ('travel_carrier', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('country_of_departure', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('departure_date', 'TEXT' if backend != 'mysql' else 'VARCHAR(64)'),
        ('arrival_date', 'TEXT' if backend != 'mysql' else 'VARCHAR(64)'),
        ('arrival_channel', 'TEXT' if backend != 'mysql' else 'VARCHAR(64)'),
        ('duration_of_stay', 'TEXT' if backend != 'mysql' else 'VARCHAR(64)'),
        ('port_of_entry', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('contact_name', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('contact_email', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('contact_phone', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('contact_address', 'TEXT'),
        ('contact_city', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('contact_state', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('contact_country', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
        ('contact_postal_code', 'TEXT' if backend != 'mysql' else 'VARCHAR(255)'),
    ]
    for column_name, column_type in required_applicant_columns:
        try:
            cursor.execute(f'ALTER TABLE applicants ADD COLUMN {column_name} {column_type}')
        except Exception:
            pass


def row_to_dict(row):
    if row is None:
        return None
    if isinstance(row, dict):
        return dict(row)
    if hasattr(row, 'keys'):
        return {k: row[k] for k in row.keys()}
    return row


def scalar_value(row, fallback=0):
    if row is None:
        return fallback
    if isinstance(row, (tuple, list)):
        return row[0] if row else fallback
    if hasattr(row, 'keys'):
        keys = list(row.keys())
        return row[keys[0]] if keys else fallback
    return row


def normalize_whitespace(value):
    return re.sub(r'\s+', ' ', str(value or '').strip())


def normalize_title(value):
    raw = normalize_whitespace(value).lower()
    return TITLE_NORMALIZATION.get(raw, normalize_whitespace(value))


def normalize_name(value):
    return normalize_whitespace(value).title()


def normalize_upper_text(value):
    return normalize_whitespace(value).upper()


def normalize_date(value):
    raw = normalize_whitespace(value)
    if not raw:
        return ''
    raw = raw.replace('.', '/').replace('-', '/')
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%d/%m/%Y', '%d %b %Y', '%d %B %Y'):
        try:
            return datetime.strptime(raw, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return raw


def nigeria_now():
    return datetime.utcnow() + timedelta(hours=1)


def parse_iso_date(value):
    normalized = normalize_date(value)
    if not normalized:
        return None
    try:
        return datetime.strptime(normalized, '%Y-%m-%d')
    except ValueError:
        return None


def normalize_display_date(value):
    normalized = normalize_date(value)
    if not normalized:
        return ''
    try:
        return datetime.strptime(normalized, '%Y-%m-%d').strftime('%d/%m/%Y')
    except ValueError:
        return normalized


def normalize_flight_number(value):
    return normalize_whitespace(value).replace(' ', '').upper()


def detect_travel_carrier(flight_number):
    normalized = normalize_flight_number(flight_number)
    if len(normalized) < 2:
        return ''
    prefix2 = normalized[:2]
    return AIRLINE_PREFIXES.get(prefix2, '')


def get_entry_ports_map(conn):
    rows = conn.execute(
        'SELECT channel, port_name FROM entry_ports WHERE is_active = 1 ORDER BY channel ASC, sort_order ASC, port_name ASC'
    ).fetchall()
    ports = {'Air': [], 'Land': [], 'Sea': []}
    for row in rows:
        record = row_to_dict(row)
        channel = record.get('channel')
        if channel in ports:
            ports[channel].append(record.get('port_name'))
    for channel, defaults in DEFAULT_ENTRY_PORTS.items():
        if not ports.get(channel):
            ports[channel] = list(defaults)
    return ports


def build_default_applicant():
    travel_dt = nigeria_now() + timedelta(days=3)
    travel_display = travel_dt.strftime('%Y-%m-%d')
    return {
        'reference_number': generate_reference(),
        'status': 'Draft',
        'passport_type': 'Standard',
        'nigerian_passport': 'No',
        'arrival_channel': 'Air',
        'marital_status': 'SINGLE',
        'gender': '',
        'departure_date': travel_display,
        'arrival_date': travel_display,
        'contact_country': 'Nigeria',
    }


def ordinal_day(day):
    if 10 <= day % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
    return f'{day}{suffix}'


def human_letter_date(value):
    normalized = normalize_date(value) or datetime.utcnow().strftime('%Y-%m-%d')
    try:
        dt = datetime.strptime(normalized, '%Y-%m-%d')
        return f"{ordinal_day(dt.day)} {dt.strftime('%B %Y')}"
    except ValueError:
        return normalized


def upload_file(file_storage, folder_name, allowed_extensions):
    ensure_upload_dirs()
    if not file_storage or not file_storage.filename:
        return None
    filename = secure_filename(file_storage.filename)
    ext = os.path.splitext(filename)[1].lower()
    if allowed_extensions and ext not in allowed_extensions:
        return None
    stamp = datetime.utcnow().strftime('%Y%m%d%H%M%S%f')
    stored_name = f'{stamp}_{filename}'
    folder = os.path.join(UPLOAD_ROOT, folder_name)
    os.makedirs(folder, exist_ok=True)
    absolute_path = os.path.join(folder, stored_name)
    file_storage.save(absolute_path)
    return {
        'absolute_path': absolute_path,
        'relative_path': to_storage_key(folder_name, stored_name),
        'stored_name': stored_name,
        'original_name': file_storage.filename,
        'size_bytes': os.path.getsize(absolute_path) if os.path.exists(absolute_path) else 0,
        'mime_type': getattr(file_storage, 'mimetype', None),
    }


def contact_label(contact):
    company = contact.get('company_name') or ''
    name = contact.get('contact_name') or ''
    return f'{name} - {company}' if company else name


def generate_reference():
    stamp = datetime.utcnow().strftime('%Y%m%d%H%M%S')
    return f'BDJ-IM-{stamp}'


def build_full_name(applicant):
    parts = [applicant.get('surname'), applicant.get('first_name'), applicant.get('middle_name')]
    return normalize_whitespace(' '.join([p for p in parts if p]))


def text_or_dash(value):
    return normalize_whitespace(value) or '-'


def normalize_yes_no(value):
    raw = normalize_whitespace(value).lower()
    if raw in ('yes', 'y', 'true', '1'):
        return 'Yes'
    if raw in ('no', 'n', 'false', '0'):
        return 'No'
    return ''


def parse_applicant_form(form):
    data = {
        'title': normalize_title(form.get('title')),
        'surname': normalize_upper_text(form.get('surname')),
        'first_name': normalize_upper_text(form.get('first_name')),
        'middle_name': normalize_upper_text(form.get('middle_name')),
        'passport_type': normalize_whitespace(form.get('passport_type')),
        'gender': normalize_whitespace(form.get('gender')),
        'marital_status': normalize_whitespace(form.get('marital_status')),
        'date_of_birth': normalize_date(form.get('date_of_birth')),
        'place_of_birth': normalize_upper_text(form.get('place_of_birth')),
        'passport_number': normalize_whitespace(form.get('passport_number')).upper(),
        'passport_expiry': normalize_date(form.get('passport_expiry')),
        'nigerian_passport': normalize_yes_no(form.get('nigerian_passport')),
        'nationality': normalize_whitespace(form.get('nationality')),
        'visa_type': normalize_whitespace(form.get('visa_type')),
        'status': normalize_whitespace(form.get('status')) or 'Pending',
        'country_of_departure': normalize_whitespace(form.get('country_of_departure')),
        'departure_date': normalize_date(form.get('departure_date')),
        'arrival_date': normalize_date(form.get('arrival_date')),
        'arrival_channel': normalize_whitespace(form.get('arrival_channel')),
        'duration_of_stay': normalize_whitespace(form.get('duration_of_stay')),
        'port_of_entry': normalize_whitespace(form.get('port_of_entry')),
        'travel_date': normalize_date(form.get('travel_date')),
        'email': normalize_whitespace(form.get('email')).lower(),
        'phone': normalize_whitespace(form.get('phone')),
        'flight_number': normalize_flight_number(form.get('flight_number')),
        'travel_carrier': normalize_whitespace(form.get('travel_carrier')),
        'reference_number': normalize_whitespace(form.get('reference_number')) or generate_reference(),
        'company_id': form.get('company_id') or None,
        'contact_id': form.get('contact_id') or None,
        'contact_name': normalize_whitespace(form.get('contact_name')),
        'contact_email': normalize_whitespace(form.get('contact_email')).lower(),
        'contact_phone': normalize_whitespace(form.get('contact_phone')),
        'contact_address': normalize_whitespace(form.get('contact_address')),
        'contact_city': normalize_whitespace(form.get('contact_city')),
        'contact_state': normalize_whitespace(form.get('contact_state')),
        'contact_country': normalize_whitespace(form.get('contact_country')),
        'contact_postal_code': normalize_whitespace(form.get('contact_postal_code')),
        'reason': normalize_whitespace(form.get('reason')),
        'notes': normalize_whitespace(form.get('notes')),
    }
    if form.get('save_mode') == 'draft':
        data['status'] = 'Draft'
    elif data['status'] == 'Draft' and form.get('save_mode') == 'final':
        data['status'] = 'Validated'
    data['travel_date'] = data['arrival_date'] or data['travel_date']
    if data['flight_number'] and not data['travel_carrier']:
        data['travel_carrier'] = detect_travel_carrier(data['flight_number'])
    data['full_name'] = build_full_name(data)
    return data


def build_letter_context(applicant, company, contact, letterhead, reason, visa_type, plain_paper):
    applicant_name = build_full_name(applicant)
    issue_date = human_letter_date(datetime.utcnow().strftime('%Y-%m-%d'))
    pronoun = 'I' if plain_paper else 'We'
    company_phrase = 'I' if plain_paper else 'Our Company'
    with_us = '' if plain_paper else ' with us'
    selected_visa = 'Visiting Visa' if reason == 'Visit' else visa_type

    body = (
        f"{issue_date}\n\n"
        f"To Whom It May Concern,\n\n"
        f"{pronoun} write to confirm that {applicant_name} ({text_or_dash(applicant.get('passport_number'))}), "
        f"a national of {text_or_dash(applicant.get('nationality'))}, has been invited by {company.get('name') or company_phrase} "
        f"for {reason.lower() if reason else 'an official visit'}{with_us}. "
        f"The applicant intends to travel on {text_or_dash(applicant.get('travel_date'))} and requires a {selected_visa}.\n\n"
        f"{company_phrase} undertake responsibility for the visit logistics in line with immigration requirements. "
        f"Please grant the necessary visa support.\n\n"
        f"Sincerely,\n"
        f"{text_or_dash(letterhead.get('signatory'))}\n"
        f"{text_or_dash(letterhead.get('designation'))}\n"
        f"{company.get('name') or ''}"
    )

    return {
        'applicant_name': applicant_name,
        'issue_date': issue_date,
        'body': body,
        'selected_visa': selected_visa,
        'reason': reason,
        'plain_paper': plain_paper,
        'contact_name': contact.get('contact_name') if contact else '',
    }


def save_audit(conn, model_id, user_id, action, entity_type, entity_id, details):
    conn.execute(
        '''
        INSERT INTO audit_logs (model_id, user_id, workspace, action, entity_type, entity_id, details)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ''',
        (model_id, user_id, IMMIGRATION_WORKSPACE, action, entity_type, entity_id, json.dumps(details or {})),
    )


def maybe_commit(conn):
    if hasattr(conn, 'commit'):
        conn.commit()


def get_company_contacts(conn, model_id, company_id=None):
    sql = '''
        SELECT contacts.*, companies.name AS company_name
        FROM contacts
        LEFT JOIN companies ON companies.id = contacts.company_id
        WHERE contacts.model_id = ?
    '''
    params = [model_id]
    if company_id:
        sql += ' AND contacts.company_id = ?'
        params.append(company_id)
    sql += ' ORDER BY contacts.contact_name ASC'
    rows = conn.execute(sql, tuple(params)).fetchall()
    return [row_to_dict(row) for row in rows]


def get_companies(conn, model_id):
    rows = conn.execute(
        'SELECT * FROM companies WHERE model_id = ? ORDER BY name ASC',
        (model_id,),
    ).fetchall()
    return [row_to_dict(row) for row in rows]


def get_letterheads(conn, model_id):
    rows = conn.execute(
        '''
        SELECT letterheads.*, companies.name AS company_name
        FROM letterheads
        LEFT JOIN companies ON companies.id = letterheads.company_id
        WHERE letterheads.model_id = ?
        ORDER BY letterheads.created_at DESC
        ''',
        (model_id,),
    ).fetchall()
    return [row_to_dict(row) for row in rows]


def get_recent_visa_letters(conn, model_id, limit=8):
    rows = conn.execute(
        '''
        SELECT visa_letters.*, applicants.full_name AS applicant_name, companies.name AS company_name
        FROM visa_letters
        LEFT JOIN applicants ON applicants.id = visa_letters.applicant_id
        LEFT JOIN companies ON companies.id = visa_letters.company_id
        WHERE visa_letters.model_id = ?
        ORDER BY visa_letters.created_at DESC
        LIMIT ?
        ''',
        (model_id, limit),
    ).fetchall()
    return [row_to_dict(row) for row in rows]


def fetch_user_preference(conn, user_id, pref_key, default=''):
    try:
        row = conn.execute(
            'SELECT pref_value FROM user_preferences WHERE user_id = ? AND pref_key = ?',
            (user_id, pref_key),
        ).fetchone()
        return scalar_value(row, default) if row else default
    except Exception:
        return default


def upsert_user_preference(conn, user_id, pref_key, pref_value):
    existing = conn.execute(
        'SELECT id FROM user_preferences WHERE user_id = ? AND pref_key = ?',
        (user_id, pref_key),
    ).fetchone()
    if existing:
        conn.execute(
            'UPDATE user_preferences SET pref_value = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?',
            (pref_value, scalar_value(existing)),
        )
    else:
        conn.execute(
            'INSERT INTO user_preferences (user_id, pref_key, pref_value) VALUES (?, ?, ?)',
            (user_id, pref_key, pref_value),
        )


def generate_report_rows(conn, model_id, filters):
    sql = '''
        SELECT applicants.*, companies.name AS company_name, contacts.contact_name
        FROM applicants
        LEFT JOIN companies ON companies.id = applicants.company_id
        LEFT JOIN contacts ON contacts.id = applicants.contact_id
        WHERE applicants.model_id = ?
    '''
    params = [model_id]
    if filters.get('visa_type'):
        sql += ' AND applicants.visa_type = ?'
        params.append(filters['visa_type'])
    if filters.get('nationality'):
        sql += ' AND applicants.nationality = ?'
        params.append(filters['nationality'])
    if filters.get('company_id'):
        sql += ' AND applicants.company_id = ?'
        params.append(filters['company_id'])
    if filters.get('travel_date'):
        sql += ' AND applicants.travel_date = ?'
        params.append(filters['travel_date'])
    sql += ' ORDER BY applicants.created_at DESC'
    rows = conn.execute(sql, tuple(params)).fetchall()
    return [row_to_dict(row) for row in rows]


def export_rows_to_csv(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Reference Number', 'Applicant', 'Passport Number', 'Nationality', 'Visa Type',
        'Status', 'Travel Date', 'Company', 'Contact', 'Email', 'Phone', 'Created At'
    ])
    for row in rows:
        writer.writerow([
            row.get('reference_number'),
            row.get('full_name'),
            row.get('passport_number'),
            row.get('nationality'),
            row.get('visa_type'),
            row.get('status'),
            row.get('travel_date'),
            row.get('company_name'),
            row.get('contact_name'),
            row.get('email'),
            row.get('phone'),
            row.get('created_at'),
        ])
    buffer = io.BytesIO(output.getvalue().encode('utf-8'))
    buffer.seek(0)
    return buffer


def export_rows_to_xlsx(rows):
    if Workbook is None:
        return None
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Applicants'
    sheet.append([
        'Reference Number', 'Applicant', 'Passport Number', 'Nationality', 'Visa Type',
        'Status', 'Travel Date', 'Company', 'Contact', 'Email', 'Phone', 'Created At'
    ])
    for row in rows:
        sheet.append([
            row.get('reference_number'),
            row.get('full_name'),
            row.get('passport_number'),
            row.get('nationality'),
            row.get('visa_type'),
            row.get('status'),
            row.get('travel_date'),
            row.get('company_name'),
            row.get('contact_name'),
            row.get('email'),
            row.get('phone'),
            str(row.get('created_at') or ''),
        ])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def export_rows_to_pdf(rows):
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 42
    pdf.setFont('Helvetica-Bold', 14)
    pdf.drawString(40, y, 'Immigration Applicants Report')
    y -= 24
    pdf.setFont('Helvetica', 9)
    for row in rows:
        line = (
            f"{row.get('reference_number', '')} | {row.get('full_name', '')} | "
            f"{row.get('passport_number', '')} | {row.get('nationality', '')} | "
            f"{row.get('visa_type', '')} | {row.get('status', '')}"
        )
        pdf.drawString(40, y, line[:140])
        y -= 16
        if y < 60:
            pdf.showPage()
            y = height - 42
            pdf.setFont('Helvetica', 9)
    pdf.save()
    buffer.seek(0)
    return buffer


def extract_passport_ocr(file_storage):
    if not file_storage or not file_storage.filename:
        return {'ok': False, 'error': 'No passport image provided.'}
    uploaded = upload_file(file_storage, 'ocr', ALLOWED_IMAGE_EXTENSIONS)
    if not uploaded:
        return {'ok': False, 'error': 'Unsupported passport image format.'}
    data = {
        'surname': '',
        'first_name': '',
        'middle_name': '',
        'passport_number': '',
        'nationality': '',
        'date_of_birth': '',
        'expiry_date': '',
        'place_of_birth': '',
    }
    message = 'OCR engine unavailable. Install pytesseract to enable automatic extraction.'
    if pytesseract is not None:
        try:
            from PIL import Image
            image = Image.open(uploaded['absolute_path'])
            raw_text = pytesseract.image_to_string(image)
            lines = [normalize_whitespace(line) for line in raw_text.splitlines() if normalize_whitespace(line)]
            joined = ' '.join(lines)
            passport_match = re.search(r'([A-Z0-9]{7,12})', joined)
            if passport_match:
                data['passport_number'] = passport_match.group(1)
            date_matches = re.findall(r'(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}[/-]\d{2}[/-]\d{2})', joined)
            if date_matches:
                data['date_of_birth'] = normalize_date(date_matches[0])
            if len(date_matches) > 1:
                data['expiry_date'] = normalize_date(date_matches[1])
            if lines:
                data['surname'] = normalize_name(lines[0].split(' ')[0])
            if len(lines) > 1:
                name_parts = lines[1].split(' ')
                if name_parts:
                    data['first_name'] = normalize_name(name_parts[0])
                    if len(name_parts) > 1:
                        data['middle_name'] = normalize_name(' '.join(name_parts[1:]))
            nationality_match = re.search(r'Nationality[:\s]+([A-Za-z ]+)', joined, re.I)
            if nationality_match:
                data['nationality'] = normalize_name(nationality_match.group(1))
            pob_match = re.search(r'Place of Birth[:\s]+([A-Za-z ]+)', joined, re.I)
            if pob_match:
                data['place_of_birth'] = normalize_name(pob_match.group(1))
            message = 'OCR extraction completed. Review and apply the values before saving.'
        except Exception as exc:
            message = f'OCR could not complete automatic extraction: {exc}'
    return {'ok': True, 'message': message, 'data': data, 'preview_path': build_file_url(uploaded['relative_path'])}


def register_immigration_routes(app, helpers):
    get_db_connection = helpers['get_db_connection']
    init_db = helpers.get('init_db')
    current_model_id = helpers['current_model_id']
    comma2 = helpers['comma2']
    set_active_workspace = helpers['set_active_workspace']
    schema_ready = {'ready': False}

    def is_missing_table_error(exc):
        message = str(exc).lower()
        return 'does not exist' in message or 'undefinedtable' in message or 'no such table' in message

    def ensure_immigration_ready(conn):
        if schema_ready['ready']:
            return
        try:
            conn.execute('SELECT 1 FROM applicants LIMIT 1')
            schema_ready['ready'] = True
            return
        except Exception as exc:
            if not is_missing_table_error(exc):
                raise
        if init_db is not None:
            init_db()
        conn.execute('SELECT 1 FROM applicants LIMIT 1')
        schema_ready['ready'] = True

    @app.context_processor
    def immigration_template_helpers():
        return {'immigration_file_url': build_file_url}

    @app.route('/immigration/files/<path:stored_path>')
    def immigration_uploaded_file(stored_path):
        absolute_path = resolve_storage_path(stored_path)
        if not absolute_path or not os.path.exists(absolute_path):
            return ('Not Found', 404)
        return send_file(absolute_path)

    @app.route('/immigration')
    @app.route('/immigration/dashboard')
    def immigration_dashboard():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        total_applicants = scalar_value(conn.execute('SELECT COUNT(*) AS value FROM applicants WHERE model_id = ?', (model_id,)).fetchone(), 0)
        applications_today = scalar_value(
            conn.execute(
                "SELECT COUNT(*) AS value FROM applicants WHERE model_id = ? AND DATE(created_at) = DATE(?)",
                (model_id, datetime.utcnow().strftime('%Y-%m-%d')),
            ).fetchone(),
            0,
        )
        pending_applications = scalar_value(
            conn.execute("SELECT COUNT(*) AS value FROM applicants WHERE model_id = ? AND status IN ('Pending', 'In Review', 'Draft')", (model_id,)).fetchone(),
            0,
        )
        completed_applications = scalar_value(
            conn.execute("SELECT COUNT(*) AS value FROM applicants WHERE model_id = ? AND status = 'Completed'", (model_id,)).fetchone(),
            0,
        )
        visa_letters_generated = scalar_value(conn.execute('SELECT COUNT(*) AS value FROM visa_letters WHERE model_id = ?', (model_id,)).fetchone(), 0)
        validation_errors = scalar_value(
            conn.execute("SELECT COUNT(*) AS value FROM documents WHERE model_id = ? AND validation_status = 'Error'", (model_id,)).fetchone(),
            0,
        )
        browser_row = row_to_dict(
            conn.execute(
                'SELECT * FROM automation_sessions WHERE model_id = ? ORDER BY updated_at DESC LIMIT 1',
                (model_id,),
            ).fetchone()
        ) or {'status': 'Disconnected', 'current_page': 'Idle', 'progress_percent': 0}

        recent_applicants = [
            row_to_dict(row)
            for row in conn.execute(
                '''
                SELECT applicants.*, companies.name AS company_name
                FROM applicants
                LEFT JOIN companies ON companies.id = applicants.company_id
                WHERE applicants.model_id = ?
                ORDER BY applicants.created_at DESC
                LIMIT 6
                ''',
                (model_id,),
            ).fetchall()
        ]
        upcoming_travel = [
            row_to_dict(row)
            for row in conn.execute(
                '''
                SELECT applicants.*, companies.name AS company_name
                FROM applicants
                LEFT JOIN companies ON companies.id = applicants.company_id
                WHERE applicants.model_id = ? AND applicants.travel_date IS NOT NULL AND applicants.travel_date != ''
                ORDER BY applicants.travel_date ASC
                LIMIT 6
                ''',
                (model_id,),
            ).fetchall()
        ]

        nationality_rows = conn.execute(
            '''
            SELECT nationality, COUNT(*) AS value
            FROM applicants
            WHERE model_id = ? AND nationality IS NOT NULL AND nationality != ''
            GROUP BY nationality
            ORDER BY value DESC
            LIMIT 8
            ''',
            (model_id,),
        ).fetchall()
        visa_type_rows = conn.execute(
            '''
            SELECT visa_type, COUNT(*) AS value
            FROM applicants
            WHERE model_id = ? AND visa_type IS NOT NULL AND visa_type != ''
            GROUP BY visa_type
            ORDER BY value DESC
            LIMIT 8
            ''',
            (model_id,),
        ).fetchall()
        company_rows = conn.execute(
            '''
            SELECT companies.name, COUNT(applicants.id) AS value
            FROM companies
            LEFT JOIN applicants ON applicants.company_id = companies.id
            WHERE companies.model_id = ?
            GROUP BY companies.id, companies.name
            ORDER BY value DESC
            LIMIT 8
            ''',
            (model_id,),
        ).fetchall()
        recent_letters = get_recent_visa_letters(conn, model_id)
        validation_summary = [
            row_to_dict(row)
            for row in conn.execute(
                '''
                SELECT validation_status, COUNT(*) AS value
                FROM documents
                WHERE model_id = ?
                GROUP BY validation_status
                ORDER BY value DESC
                ''',
                (model_id,),
            ).fetchall()
        ]

        return render_template(
            'immigration/dashboard.html',
            page_title='Immigration Dashboard',
            stats={
                'total_applicants': total_applicants,
                'applications_today': applications_today,
                'pending_applications': pending_applications,
                'completed_applications': completed_applications,
                'visa_letters_generated': visa_letters_generated,
                'validation_errors': validation_errors,
                'browser_status': browser_row.get('status', 'Disconnected'),
            },
            browser_row=browser_row,
            recent_applicants=recent_applicants,
            upcoming_travel=upcoming_travel,
            nationality_chart=[row_to_dict(row) for row in nationality_rows],
            visa_type_chart=[row_to_dict(row) for row in visa_type_rows],
            company_activity=[row_to_dict(row) for row in company_rows],
            recent_letters=recent_letters,
            validation_summary=validation_summary,
        )

    @app.route('/immigration/applicants')
    def immigration_applicants():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        search = normalize_whitespace(request.args.get('q'))
        sort = normalize_whitespace(request.args.get('sort')) or 'newest'
        company_id = request.args.get('company_id') or ''
        contact_id = request.args.get('contact_id') or ''
        nationality = normalize_whitespace(request.args.get('nationality'))
        visa_type = normalize_whitespace(request.args.get('visa_type'))
        status = normalize_whitespace(request.args.get('status'))
        travel_date = normalize_date(request.args.get('travel_date'))
        created_date = normalize_date(request.args.get('created_date'))

        sql = '''
            SELECT applicants.*, companies.name AS company_name, contacts.contact_name
            FROM applicants
            LEFT JOIN companies ON companies.id = applicants.company_id
            LEFT JOIN contacts ON contacts.id = applicants.contact_id
            WHERE applicants.model_id = ?
        '''
        params = [model_id]
        if search:
            like = f'%{search}%'
            sql += '''
                AND (
                    applicants.full_name LIKE ? OR
                    applicants.surname LIKE ? OR
                    applicants.first_name LIKE ? OR
                    applicants.passport_number LIKE ? OR
                    applicants.nationality LIKE ? OR
                    applicants.visa_type LIKE ? OR
                    applicants.email LIKE ? OR
                    applicants.reference_number LIKE ? OR
                    contacts.contact_name LIKE ? OR
                    companies.name LIKE ?
                )
            '''
            params.extend([like] * 10)
        if company_id:
            sql += ' AND applicants.company_id = ?'
            params.append(company_id)
        if contact_id:
            sql += ' AND applicants.contact_id = ?'
            params.append(contact_id)
        if nationality:
            sql += ' AND applicants.nationality = ?'
            params.append(nationality)
        if visa_type:
            sql += ' AND applicants.visa_type = ?'
            params.append(visa_type)
        if status:
            sql += ' AND applicants.status = ?'
            params.append(status)
        if travel_date:
            sql += ' AND applicants.travel_date = ?'
            params.append(travel_date)
        if created_date:
            sql += ' AND DATE(applicants.created_at) = DATE(?)'
            params.append(created_date)

        sort_map = {
            'newest': 'applicants.created_at DESC',
            'oldest': 'applicants.created_at ASC',
            'name_asc': 'applicants.full_name ASC',
            'name_desc': 'applicants.full_name DESC',
            'travel_date': 'applicants.travel_date ASC',
            'visa_type': 'applicants.visa_type ASC',
        }
        sql += f' ORDER BY {sort_map.get(sort, sort_map["newest"])}'
        rows = conn.execute(sql, tuple(params)).fetchall()
        applicants = [row_to_dict(row) for row in rows]

        companies = get_companies(conn, model_id)
        contacts = get_company_contacts(conn, model_id, company_id if company_id else None)
        nationalities = sorted({(row.get('nationality') or '').strip() for row in applicants if row.get('nationality')})

        return render_template(
            'immigration/applicants.html',
            page_title='Applicants',
            applicants=applicants,
            companies=companies,
            contacts=contacts,
            nationalities=nationalities,
            visa_types=VISA_CATEGORIES,
            statuses=APPLICANT_STATUSES,
            reasons=REASONS,
            filters={
                'q': search,
                'sort': sort,
                'company_id': company_id,
                'contact_id': contact_id,
                'nationality': nationality,
                'visa_type': visa_type,
                'status': status,
                'travel_date': travel_date,
                'created_date': created_date,
            },
        )

    @app.route('/immigration/applicants/new')
    @app.route('/immigration/applicants/<int:applicant_id>/edit')
    def immigration_applicant_form(applicant_id=None):
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        applicant = build_default_applicant()
        if applicant_id:
            applicant = row_to_dict(
                conn.execute('SELECT * FROM applicants WHERE id = ? AND model_id = ?', (applicant_id, model_id)).fetchone()
            )
            if not applicant:
                return redirect(url_for('immigration_applicants'))
        else:
            applicant['company_id'] = None
            applicant['contact_id'] = None
        companies = get_companies(conn, model_id)
        contacts = get_company_contacts(conn, model_id, applicant.get('company_id') if applicant else None)
        documents = []
        if applicant_id:
            documents = [
                row_to_dict(row)
                for row in conn.execute(
                    'SELECT * FROM documents WHERE applicant_id = ? ORDER BY created_at DESC',
                    (applicant_id,),
                ).fetchall()
            ]
        return render_template(
            'immigration/applicant_form.html',
            page_title='Edit Applicant' if applicant_id else 'Add Applicant',
            applicant=applicant,
            companies=companies,
            contacts=contacts,
            all_contacts=get_company_contacts(conn, model_id),
            applicant_titles=APPLICANT_TITLES,
            passport_types=PASSPORT_TYPES,
            gender_options=GENDER_OPTIONS,
            marital_status_options=MARITAL_STATUS_OPTIONS,
            visa_categories=VISA_CATEGORIES,
            journey_purposes=JOURNEY_PURPOSES,
            reasons=REASONS,
            statuses=APPLICANT_STATUSES,
            arrival_channels=ARRIVAL_CHANNELS,
            ports_by_channel=get_entry_ports_map(conn),
            travel_carriers=TRAVEL_CARRIERS,
            documents=documents,
            nationality_options=NATIONALITY_OPTIONS,
            country_options=COUNTRY_OPTIONS,
            success_message=request.args.get('message', ''),
        )

    @app.route('/immigration/applicants/save', methods=['POST'])
    @app.route('/immigration/applicants/<int:applicant_id>/save', methods=['POST'])
    def save_immigration_applicant(applicant_id=None):
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        payload = parse_applicant_form(request.form)

        def render_form_error(message):
            return render_template(
                'immigration/applicant_form.html',
                page_title='Applicant Form',
                error=message,
                applicant=payload,
                companies=get_companies(conn, model_id),
                contacts=get_company_contacts(conn, model_id, payload.get('company_id')),
                all_contacts=get_company_contacts(conn, model_id),
                applicant_titles=APPLICANT_TITLES,
                passport_types=PASSPORT_TYPES,
                gender_options=GENDER_OPTIONS,
                marital_status_options=MARITAL_STATUS_OPTIONS,
                visa_categories=VISA_CATEGORIES,
                journey_purposes=JOURNEY_PURPOSES,
                reasons=REASONS,
                statuses=APPLICANT_STATUSES,
                arrival_channels=ARRIVAL_CHANNELS,
                ports_by_channel=get_entry_ports_map(conn),
                travel_carriers=TRAVEL_CARRIERS,
                documents=[],
                nationality_options=NATIONALITY_OPTIONS,
                country_options=COUNTRY_OPTIONS,
                success_message='',
            )

        required_fields = [
            ('nationality', 'Nationality is required.'),
            ('visa_type', 'Visa Category is required.'),
            ('passport_type', 'Passport Type is required.'),
            ('title', 'Title is required.'),
            ('surname', 'Surname is required.'),
            ('first_name', 'First Name is required.'),
            ('date_of_birth', 'Date of Birth is required.'),
            ('place_of_birth', 'Place of Birth is required.'),
            ('gender', 'Gender is required.'),
            ('marital_status', 'Marital Status is required.'),
            ('passport_number', 'Passport Number is required.'),
            ('passport_expiry', 'Passport Expiry Date is required.'),
            ('nigerian_passport', 'Nigerian Passport selection is required.'),
            ('reason', 'Purpose of Journey is required.'),
            ('travel_carrier', 'Travel Carrier is required.'),
            ('flight_number', 'Flight Number is required.'),
            ('country_of_departure', 'Country of Departure is required.'),
            ('departure_date', 'Departure Date is required.'),
            ('arrival_date', 'Arrival Date is required.'),
            ('arrival_channel', 'Arrival Channel is required.'),
            ('duration_of_stay', 'Duration of Stay is required.'),
            ('port_of_entry', 'Port of Entry is required.'),
            ('contact_name', 'Contact Name is required.'),
            ('contact_phone', 'Contact Phone is required.'),
            ('contact_address', 'Contact Address is required.'),
            ('contact_city', 'Contact City is required.'),
            ('contact_state', 'Contact State is required.'),
            ('contact_email', 'Contact Email is required.'),
            ('contact_country', 'Contact Country is required.'),
        ]
        for field_name, message in required_fields:
            if not payload.get(field_name):
                return render_form_error(message)

        dob_dt = parse_iso_date(payload.get('date_of_birth'))
        passport_expiry_dt = parse_iso_date(payload.get('passport_expiry'))
        departure_dt = parse_iso_date(payload.get('departure_date'))
        arrival_dt = parse_iso_date(payload.get('arrival_date'))
        today = nigeria_now().date()

        if dob_dt and dob_dt.date() > today:
            return render_form_error('Date of Birth cannot be in the future.')
        if passport_expiry_dt and passport_expiry_dt.date() <= today:
            return render_form_error('Passport Expiry Date must be later than today.')
        if departure_dt and arrival_dt and arrival_dt.date() < departure_dt.date():
            return render_form_error('Arrival Date cannot be earlier than Departure Date.')

        duplicate_row = conn.execute(
            'SELECT id FROM applicants WHERE model_id = ? AND passport_number = ? AND (? IS NULL OR id != ?)',
            (model_id, payload['passport_number'], applicant_id, applicant_id),
        ).fetchone()
        if duplicate_row:
            return render_form_error('An applicant with this Passport Number already exists.')

        values = (
            model_id,
            payload['title'],
            payload['surname'],
            payload['first_name'],
            payload['middle_name'],
            payload['full_name'],
            payload['passport_type'],
            payload['gender'],
            payload['marital_status'],
            payload['date_of_birth'],
            payload['place_of_birth'],
            payload['passport_number'],
            payload['passport_expiry'],
            payload['nigerian_passport'],
            payload['flight_number'],
            payload['travel_carrier'],
            payload['nationality'],
            payload['visa_type'],
            payload['status'],
            payload['country_of_departure'],
            payload['departure_date'],
            payload['arrival_date'],
            payload['arrival_channel'],
            payload['duration_of_stay'],
            payload['port_of_entry'],
            payload['travel_date'],
            payload['email'],
            payload['phone'],
            payload['contact_name'],
            payload['contact_email'],
            payload['contact_phone'],
            payload['contact_address'],
            payload['contact_city'],
            payload['contact_state'],
            payload['contact_country'],
            payload['contact_postal_code'],
            payload['reference_number'],
            payload['company_id'],
            payload['contact_id'],
            payload['reason'],
            payload['notes'],
            json.dumps(payload),
            session.get('username'),
        )
        if applicant_id:
            conn.execute(
                '''
                UPDATE applicants SET
                    model_id = ?, title = ?, surname = ?, first_name = ?, middle_name = ?, full_name = ?,
                    passport_type = ?, gender = ?, marital_status = ?, date_of_birth = ?, place_of_birth = ?, passport_number = ?, passport_expiry = ?,
                    nigerian_passport = ?, flight_number = ?, travel_carrier = ?, nationality = ?, visa_type = ?, status = ?,
                    country_of_departure = ?, departure_date = ?, arrival_date = ?, arrival_channel = ?, duration_of_stay = ?, port_of_entry = ?, travel_date = ?, email = ?, phone = ?,
                    contact_name = ?, contact_email = ?, contact_phone = ?, contact_address = ?, contact_city = ?,
                    contact_state = ?, contact_country = ?, contact_postal_code = ?, reference_number = ?, company_id = ?, contact_id = ?, reason = ?, notes = ?,
                    draft_data = ?, created_by = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND model_id = ?
                ''',
                values + (applicant_id, model_id),
            )
            save_audit(conn, model_id, session.get('user_id'), 'update', 'applicant', applicant_id, payload)
        else:
            conn.execute(
                '''
                INSERT INTO applicants (
                    model_id, title, surname, first_name, middle_name, full_name, passport_type, gender, marital_status, date_of_birth, place_of_birth,
                    passport_number, passport_expiry, nigerian_passport, flight_number, travel_carrier,
                    nationality, visa_type, status, country_of_departure, departure_date, arrival_date, arrival_channel, duration_of_stay, port_of_entry, travel_date, email, phone,
                    contact_name, contact_email, contact_phone, contact_address, contact_city, contact_state, contact_country, contact_postal_code,
                    reference_number, company_id, contact_id, reason, notes, draft_data, created_by
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                values,
            )
            inserted = conn.execute(
                'SELECT id FROM applicants WHERE model_id = ? AND reference_number = ? ORDER BY id DESC LIMIT 1',
                (model_id, payload['reference_number']),
            ).fetchone()
            inserted_id = scalar_value(inserted)
            save_audit(conn, model_id, session.get('user_id'), 'create', 'applicant', inserted_id, payload)
        maybe_commit(conn)
        if request.form.get('next_action') == 'visa_letter':
            target_id = applicant_id or inserted_id
            return redirect(url_for('immigration_visa_letters', applicant_id=target_id))
        target_id = applicant_id or inserted_id
        return redirect(url_for('immigration_applicant_form', applicant_id=target_id, message='Applicant saved successfully.'))

    @app.route('/immigration/applicants/<int:applicant_id>/duplicate', methods=['POST'])
    def duplicate_immigration_applicant(applicant_id):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        source = row_to_dict(conn.execute('SELECT * FROM applicants WHERE id = ? AND model_id = ?', (applicant_id, model_id)).fetchone())
        if not source:
            return redirect(url_for('immigration_applicants'))
        new_ref = f"{source.get('reference_number') or generate_reference()}-COPY"
        conn.execute(
            '''
            INSERT INTO applicants (
                model_id, title, surname, first_name, middle_name, full_name, passport_type, gender, marital_status, date_of_birth, place_of_birth,
                passport_number, passport_expiry, nigerian_passport, flight_number, travel_carrier,
                nationality, visa_type, status, country_of_departure, departure_date, arrival_date, arrival_channel, duration_of_stay, port_of_entry, travel_date, email, phone,
                contact_name, contact_email, contact_phone, contact_address, contact_city, contact_state, contact_country, contact_postal_code,
                reference_number, company_id, contact_id, reason, notes, draft_data, created_by
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                model_id, source.get('title'), source.get('surname'), source.get('first_name'), source.get('middle_name'),
                source.get('full_name'), source.get('passport_type'), source.get('gender'), source.get('marital_status'),
                source.get('date_of_birth'), source.get('place_of_birth'), source.get('passport_number'), source.get('passport_expiry'),
                source.get('nigerian_passport'), source.get('flight_number'), source.get('travel_carrier'), source.get('nationality'),
                source.get('visa_type'), 'Draft', source.get('country_of_departure'), source.get('departure_date'),
                source.get('arrival_date'), source.get('arrival_channel'), source.get('duration_of_stay'), source.get('port_of_entry'),
                source.get('travel_date'), source.get('email'), source.get('phone'),
                source.get('contact_name'), source.get('contact_email'), source.get('contact_phone'), source.get('contact_address'),
                source.get('contact_city'), source.get('contact_state'), source.get('contact_country'), source.get('contact_postal_code'),
                new_ref, source.get('company_id'), source.get('contact_id'), source.get('reason'),
                source.get('notes'), source.get('draft_data'), session.get('username')
            ),
        )
        inserted = conn.execute(
            'SELECT id FROM applicants WHERE model_id = ? AND reference_number = ? ORDER BY id DESC LIMIT 1',
            (model_id, new_ref),
        ).fetchone()
        save_audit(conn, model_id, session.get('user_id'), 'duplicate', 'applicant', scalar_value(inserted), {'source_id': applicant_id})
        maybe_commit(conn)
        return redirect(url_for('immigration_applicants'))

    @app.route('/immigration/applicants/<int:applicant_id>/delete', methods=['POST'])
    def delete_immigration_applicant(applicant_id):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        conn.execute('DELETE FROM documents WHERE applicant_id = ? AND model_id = ?', (applicant_id, model_id))
        conn.execute('DELETE FROM applicants WHERE id = ? AND model_id = ?', (applicant_id, model_id))
        save_audit(conn, model_id, session.get('user_id'), 'delete', 'applicant', applicant_id, {})
        maybe_commit(conn)
        return redirect(url_for('immigration_applicants'))

    @app.route('/immigration/applicants/import', methods=['POST'])
    def import_immigration_applicants():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        upload = request.files.get('file')
        if not upload or not upload.filename:
            return redirect(url_for('immigration_applicants'))

        ext = os.path.splitext(upload.filename)[1].lower()
        rows = []
        if ext in ('.xlsx', '.xls') and load_workbook is not None:
            workbook = load_workbook(upload, read_only=True)
            sheet = workbook.active
            headers = []
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if idx == 0:
                    headers = [normalize_whitespace(cell).lower().replace(' ', '_') for cell in row]
                    continue
                rows.append(dict(zip(headers, row)))
        else:
            stream = io.StringIO(upload.stream.read().decode('utf-8'))
            reader = csv.DictReader(stream)
            rows = [dict(row) for row in reader]

        for row in rows:
            surname = normalize_name(row.get('surname'))
            first_name = normalize_name(row.get('first_name'))
            if not surname or not first_name:
                continue
            payload = {
                'title': normalize_title(row.get('title')),
                'surname': surname,
                'first_name': first_name,
                'middle_name': normalize_name(row.get('middle_name')),
                'full_name': normalize_whitespace(f"{surname} {first_name} {normalize_name(row.get('middle_name'))}"),
                'passport_type': normalize_whitespace(row.get('passport_type')),
                'gender': normalize_whitespace(row.get('gender')),
                'marital_status': normalize_whitespace(row.get('marital_status')),
                'date_of_birth': normalize_date(row.get('date_of_birth')),
                'place_of_birth': normalize_name(row.get('place_of_birth')),
                'passport_number': normalize_whitespace(row.get('passport_number')).upper(),
                'passport_expiry': normalize_date(row.get('passport_expiry')),
                'nigerian_passport': normalize_yes_no(row.get('nigerian_passport')),
                'flight_number': normalize_flight_number(row.get('flight_number')),
                'travel_carrier': normalize_whitespace(row.get('travel_carrier')) or detect_travel_carrier(row.get('flight_number')),
                'nationality': normalize_name(row.get('nationality')),
                'visa_type': normalize_whitespace(row.get('visa_type')) or 'Business Visa',
                'status': normalize_whitespace(row.get('status')) or 'Pending',
                'country_of_departure': normalize_name(row.get('country_of_departure')),
                'departure_date': normalize_date(row.get('departure_date')),
                'arrival_date': normalize_date(row.get('arrival_date')),
                'arrival_channel': normalize_whitespace(row.get('arrival_channel')),
                'duration_of_stay': normalize_whitespace(row.get('duration_of_stay')),
                'port_of_entry': normalize_whitespace(row.get('port_of_entry')),
                'travel_date': normalize_date(row.get('travel_date')),
                'email': normalize_whitespace(row.get('email')).lower(),
                'phone': normalize_whitespace(row.get('phone')),
                'contact_name': normalize_name(row.get('contact_name')),
                'contact_email': normalize_whitespace(row.get('contact_email')).lower(),
                'contact_phone': normalize_whitespace(row.get('contact_phone')),
                'contact_address': normalize_whitespace(row.get('contact_address')),
                'contact_city': normalize_name(row.get('contact_city')),
                'contact_state': normalize_name(row.get('contact_state')),
                'contact_country': normalize_name(row.get('contact_country')),
                'contact_postal_code': normalize_whitespace(row.get('contact_postal_code')),
                'reference_number': normalize_whitespace(row.get('reference_number')) or generate_reference(),
                'company_id': None,
                'contact_id': None,
                'reason': normalize_whitespace(row.get('reason')) or 'Business Meeting',
                'notes': normalize_whitespace(row.get('notes')),
            }
            conn.execute(
                '''
                INSERT INTO applicants (
                    model_id, title, surname, first_name, middle_name, full_name, passport_type, gender, marital_status, date_of_birth, place_of_birth,
                    passport_number, passport_expiry, nigerian_passport, flight_number, travel_carrier, nationality, visa_type, status,
                    country_of_departure, departure_date, arrival_date, arrival_channel, duration_of_stay, port_of_entry, travel_date, email, phone,
                    contact_name, contact_email, contact_phone, contact_address, contact_city, contact_state, contact_country, contact_postal_code,
                    reference_number, company_id, contact_id, reason, notes, draft_data, created_by
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    model_id, payload['title'], payload['surname'], payload['first_name'], payload['middle_name'],
                    payload['full_name'], payload['passport_type'], payload['gender'], payload['marital_status'], payload['date_of_birth'], payload['place_of_birth'],
                    payload['passport_number'], payload['passport_expiry'], payload['nigerian_passport'], payload['flight_number'], payload['travel_carrier'],
                    payload['nationality'], payload['visa_type'], payload['status'], payload['country_of_departure'], payload['departure_date'],
                    payload['arrival_date'], payload['arrival_channel'], payload['duration_of_stay'], payload['port_of_entry'], payload['travel_date'],
                    payload['email'], payload['phone'], payload['contact_name'], payload['contact_email'],
                    payload['contact_phone'], payload['contact_address'], payload['contact_city'], payload['contact_state'],
                    payload['contact_country'], payload['contact_postal_code'], payload['reference_number'],
                    payload['company_id'], payload['contact_id'], payload['reason'], payload['notes'],
                    json.dumps(payload), session.get('username'),
                ),
            )
        maybe_commit(conn)
        return redirect(url_for('immigration_applicants'))

    @app.route('/immigration/applicants/export')
    def export_immigration_applicants():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        rows = generate_report_rows(conn, current_model_id(), request.args)
        format_type = normalize_whitespace(request.args.get('format')) or 'csv'
        if format_type == 'xlsx':
            buffer = export_rows_to_xlsx(rows)
            if buffer is not None:
                return send_file(
                    buffer,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='immigration_applicants.xlsx',
                )
        if format_type == 'pdf':
            return send_file(
                export_rows_to_pdf(rows),
                mimetype='application/pdf',
                as_attachment=True,
                download_name='immigration_applicants.pdf',
            )
        return send_file(
            export_rows_to_csv(rows),
            mimetype='text/csv',
            as_attachment=True,
            download_name='immigration_applicants.csv',
        )

    @app.route('/immigration/applicants/ocr', methods=['POST'])
    def ocr_immigration_passport():
        result = extract_passport_ocr(request.files.get('passport_image'))
        return jsonify(result)

    @app.route('/immigration/contacts')
    def immigration_contacts():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        edit_id = request.args.get('edit_id')
        rows = get_company_contacts(conn, model_id)
        form_contact = None
        if edit_id:
            form_contact = row_to_dict(conn.execute('SELECT * FROM contacts WHERE id = ? AND model_id = ?', (edit_id, model_id)).fetchone())
        return render_template(
            'immigration/contacts.html',
            page_title='Contacts',
            contacts=rows,
            companies=get_companies(conn, model_id),
            form_contact=form_contact,
            countries=DEFAULT_COUNTRIES,
        )

    @app.route('/immigration/contacts/save', methods=['POST'])
    @app.route('/immigration/contacts/<int:contact_id>/save', methods=['POST'])
    def save_immigration_contact(contact_id=None):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        payload = {
            'company_id': request.form.get('company_id') or None,
            'contact_name': normalize_name(request.form.get('contact_name')),
            'phone': normalize_whitespace(request.form.get('phone')),
            'email': normalize_whitespace(request.form.get('email')).lower(),
            'address': normalize_whitespace(request.form.get('address')),
            'city': normalize_name(request.form.get('city')),
            'state': normalize_name(request.form.get('state')),
            'postal_code': normalize_whitespace(request.form.get('postal_code')),
            'country': normalize_name(request.form.get('country')),
        }
        if not payload['contact_name']:
            return redirect(url_for('immigration_contacts'))
        if contact_id:
            conn.execute(
                '''
                UPDATE contacts SET
                    company_id = ?, contact_name = ?, phone = ?, email = ?, address = ?,
                    city = ?, state = ?, postal_code = ?, country = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND model_id = ?
                ''',
                (
                    payload['company_id'], payload['contact_name'], payload['phone'], payload['email'],
                    payload['address'], payload['city'], payload['state'], payload['postal_code'], payload['country'],
                    contact_id, model_id,
                ),
            )
            save_audit(conn, model_id, session.get('user_id'), 'update', 'contact', contact_id, payload)
        else:
            conn.execute(
                '''
                INSERT INTO contacts (model_id, company_id, contact_name, phone, email, address, city, state, postal_code, country)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    model_id, payload['company_id'], payload['contact_name'], payload['phone'], payload['email'],
                    payload['address'], payload['city'], payload['state'], payload['postal_code'], payload['country'],
                ),
            )
        maybe_commit(conn)
        return redirect(url_for('immigration_contacts'))

    @app.route('/immigration/contacts/<int:contact_id>/duplicate', methods=['POST'])
    def duplicate_immigration_contact(contact_id):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        row = row_to_dict(conn.execute('SELECT * FROM contacts WHERE id = ? AND model_id = ?', (contact_id, model_id)).fetchone())
        if row:
            conn.execute(
                '''
                INSERT INTO contacts (model_id, company_id, contact_name, phone, email, address, city, state, postal_code, country)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    model_id, row.get('company_id'), f"{row.get('contact_name')} Copy", row.get('phone'), row.get('email'),
                    row.get('address'), row.get('city'), row.get('state'), row.get('postal_code'), row.get('country'),
                ),
            )
            maybe_commit(conn)
        return redirect(url_for('immigration_contacts'))

    @app.route('/immigration/contacts/<int:contact_id>/delete', methods=['POST'])
    def delete_immigration_contact(contact_id):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        conn.execute('DELETE FROM contacts WHERE id = ? AND model_id = ?', (contact_id, model_id))
        maybe_commit(conn)
        return redirect(url_for('immigration_contacts'))

    @app.route('/immigration/contacts/quick-add', methods=['POST'])
    def quick_add_immigration_contact():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        payload = {
            'company_id': request.form.get('company_id') or None,
            'contact_name': normalize_name(request.form.get('contact_name')),
            'phone': normalize_whitespace(request.form.get('phone')),
            'email': normalize_whitespace(request.form.get('email')).lower(),
            'address': normalize_whitespace(request.form.get('address')),
            'city': normalize_name(request.form.get('city')),
            'state': normalize_name(request.form.get('state')),
            'postal_code': normalize_whitespace(request.form.get('postal_code')),
            'country': normalize_name(request.form.get('country')),
        }
        if not payload['contact_name']:
            return jsonify({'ok': False, 'error': 'Contact Name is required.'}), 400
        conn.execute(
            '''
            INSERT INTO contacts (model_id, company_id, contact_name, phone, email, address, city, state, postal_code, country)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                model_id, payload['company_id'], payload['contact_name'], payload['phone'], payload['email'],
                payload['address'], payload['city'], payload['state'], payload['postal_code'], payload['country'],
            ),
        )
        maybe_commit(conn)
        created = row_to_dict(
            conn.execute(
                '''
                SELECT contacts.*, companies.name AS company_name
                FROM contacts
                LEFT JOIN companies ON companies.id = contacts.company_id
                WHERE contacts.model_id = ? AND contacts.contact_name = ?
                ORDER BY contacts.id DESC
                LIMIT 1
                ''',
                (model_id, payload['contact_name']),
            ).fetchone()
        )
        return jsonify({'ok': True, 'contact': created, 'label': contact_label(created)})

    @app.route('/api/immigration/contacts')
    def immigration_contacts_api():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        company_id = request.args.get('company_id') or None
        contacts = get_company_contacts(conn, model_id, company_id)
        return jsonify({'contacts': contacts})

    @app.route('/immigration/companies')
    def immigration_companies():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        edit_id = request.args.get('edit_id')
        companies = get_companies(conn, model_id)
        form_company = None
        if edit_id:
            form_company = row_to_dict(conn.execute('SELECT * FROM companies WHERE id = ? AND model_id = ?', (edit_id, model_id)).fetchone())
        return render_template(
            'immigration/companies.html',
            page_title='Companies',
            companies=companies,
            form_company=form_company,
            countries=DEFAULT_COUNTRIES,
        )

    @app.route('/immigration/companies/save', methods=['POST'])
    @app.route('/immigration/companies/<int:company_id>/save', methods=['POST'])
    def save_immigration_company(company_id=None):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        logo = upload_file(request.files.get('logo'), 'companies', ALLOWED_IMAGE_EXTENSIONS)
        payload = {
            'name': normalize_name(request.form.get('name')),
            'logo_path': logo['relative_path'] if logo else request.form.get('existing_logo_path'),
            'email': normalize_whitespace(request.form.get('email')).lower(),
            'phone': normalize_whitespace(request.form.get('phone')),
            'address': normalize_whitespace(request.form.get('address')),
            'country': normalize_name(request.form.get('country')),
            'default_contact_id': request.form.get('default_contact_id') or None,
            'default_letterhead_id': request.form.get('default_letterhead_id') or None,
        }
        if not payload['name']:
            return redirect(url_for('immigration_companies'))
        if company_id:
            conn.execute(
                '''
                UPDATE companies SET
                    name = ?, logo_path = ?, email = ?, phone = ?, address = ?, country = ?,
                    default_contact_id = ?, default_letterhead_id = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND model_id = ?
                ''',
                (
                    payload['name'], payload['logo_path'], payload['email'], payload['phone'], payload['address'],
                    payload['country'], payload['default_contact_id'], payload['default_letterhead_id'], company_id, model_id,
                ),
            )
        else:
            conn.execute(
                '''
                INSERT INTO companies (model_id, name, logo_path, email, phone, address, country, default_contact_id, default_letterhead_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    model_id, payload['name'], payload['logo_path'], payload['email'], payload['phone'],
                    payload['address'], payload['country'], payload['default_contact_id'], payload['default_letterhead_id'],
                ),
            )
        maybe_commit(conn)
        return redirect(url_for('immigration_companies'))

    @app.route('/immigration/companies/<int:company_id>/delete', methods=['POST'])
    def delete_immigration_company(company_id):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        conn.execute('DELETE FROM companies WHERE id = ? AND model_id = ?', (company_id, model_id))
        maybe_commit(conn)
        return redirect(url_for('immigration_companies'))

    @app.route('/immigration/letterheads')
    def immigration_letterheads():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        edit_id = request.args.get('edit_id')
        form_letterhead = None
        if edit_id:
            form_letterhead = row_to_dict(conn.execute('SELECT * FROM letterheads WHERE id = ? AND model_id = ?', (edit_id, model_id)).fetchone())
        return render_template(
            'immigration/letterheads.html',
            page_title='Letterheads',
            companies=get_companies(conn, model_id),
            letterheads=get_letterheads(conn, model_id),
            form_letterhead=form_letterhead,
        )

    @app.route('/immigration/letterheads/save', methods=['POST'])
    @app.route('/immigration/letterheads/<int:letterhead_id>/save', methods=['POST'])
    def save_immigration_letterhead(letterhead_id=None):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        background = upload_file(request.files.get('background_image'), 'letterheads', ALLOWED_IMAGE_EXTENSIONS)
        signature = upload_file(request.files.get('signature_image'), 'signatures', ALLOWED_IMAGE_EXTENSIONS)
        payload = {
            'company_id': request.form.get('company_id') or None,
            'template_name': normalize_whitespace(request.form.get('template_name')),
            'template_type': normalize_whitespace(request.form.get('template_type')),
            'background_image_path': background['relative_path'] if background else request.form.get('existing_background_image_path'),
            'signature_image_path': signature['relative_path'] if signature else request.form.get('existing_signature_image_path'),
            'signatory': normalize_name(request.form.get('signatory')),
            'designation': normalize_name(request.form.get('designation')),
            'scale_percent': request.form.get('scale_percent') or 100,
            'move_left': request.form.get('move_left') or 0,
            'move_right': request.form.get('move_right') or 0,
            'move_up': request.form.get('move_up') or 0,
            'move_down': request.form.get('move_down') or 0,
            'rotation': request.form.get('rotation') or 0,
            'plain_paper': 1 if request.form.get('plain_paper') else 0,
        }
        if not payload['template_name']:
            return redirect(url_for('immigration_letterheads'))
        if letterhead_id:
            conn.execute(
                '''
                UPDATE letterheads SET
                    company_id = ?, template_name = ?, template_type = ?, background_image_path = ?, signature_image_path = ?,
                    signatory = ?, designation = ?, scale_percent = ?, move_left = ?, move_right = ?, move_up = ?,
                    move_down = ?, rotation = ?, plain_paper = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND model_id = ?
                ''',
                (
                    payload['company_id'], payload['template_name'], payload['template_type'], payload['background_image_path'],
                    payload['signature_image_path'], payload['signatory'], payload['designation'], payload['scale_percent'],
                    payload['move_left'], payload['move_right'], payload['move_up'], payload['move_down'], payload['rotation'],
                    payload['plain_paper'], letterhead_id, model_id,
                ),
            )
        else:
            conn.execute(
                '''
                INSERT INTO letterheads (
                    model_id, company_id, template_name, template_type, background_image_path, signature_image_path,
                    signatory, designation, scale_percent, move_left, move_right, move_up, move_down, rotation, plain_paper
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (
                    model_id, payload['company_id'], payload['template_name'], payload['template_type'], payload['background_image_path'],
                    payload['signature_image_path'], payload['signatory'], payload['designation'], payload['scale_percent'],
                    payload['move_left'], payload['move_right'], payload['move_up'], payload['move_down'], payload['rotation'],
                    payload['plain_paper'],
                ),
            )
        maybe_commit(conn)
        return redirect(url_for('immigration_letterheads'))

    @app.route('/immigration/letterheads/<int:letterhead_id>/delete', methods=['POST'])
    def delete_immigration_letterhead(letterhead_id):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        conn.execute('DELETE FROM letterheads WHERE id = ? AND model_id = ?', (letterhead_id, model_id))
        maybe_commit(conn)
        return redirect(url_for('immigration_letterheads'))

    @app.route('/immigration/documents')
    def immigration_documents():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        applicant_id = request.args.get('applicant_id')
        applicants = [
            row_to_dict(row)
            for row in conn.execute(
                'SELECT id, full_name, reference_number FROM applicants WHERE model_id = ? ORDER BY full_name ASC',
                (model_id,),
            ).fetchall()
        ]
        sql = '''
            SELECT documents.*, applicants.full_name AS applicant_name, applicants.reference_number
            FROM documents
            LEFT JOIN applicants ON applicants.id = documents.applicant_id
            WHERE documents.model_id = ?
        '''
        params = [model_id]
        if applicant_id:
            sql += ' AND documents.applicant_id = ?'
            params.append(applicant_id)
        sql += ' ORDER BY documents.created_at DESC'
        docs = [row_to_dict(row) for row in conn.execute(sql, tuple(params)).fetchall()]
        return render_template(
            'immigration/documents.html',
            page_title='Documents',
            documents=docs,
            applicants=applicants,
            selected_applicant_id=applicant_id,
            document_types=DOCUMENT_TYPES,
        )

    @app.route('/immigration/documents/upload', methods=['POST'])
    def upload_immigration_document():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        applicant_id = request.form.get('applicant_id')
        document_type = normalize_whitespace(request.form.get('document_type')) or 'Additional Documents'
        upload = upload_file(request.files.get('document'), 'documents', ALLOWED_DOCUMENT_EXTENSIONS)
        if not applicant_id or not upload:
            return redirect(url_for('immigration_documents', applicant_id=applicant_id))
        validation_status = 'Validated' if document_type in ('Passport', 'Photograph', 'Invitation') else 'Pending'
        conn.execute(
            '''
            INSERT INTO documents (
                model_id, applicant_id, document_type, original_name, stored_name, file_path,
                mime_type, size_bytes, validation_status, extracted_data
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                model_id, applicant_id, document_type, upload['original_name'], upload['stored_name'], upload['relative_path'],
                upload['mime_type'], upload['size_bytes'], validation_status, request.form.get('extracted_data') or '',
            ),
        )
        maybe_commit(conn)
        return redirect(url_for('immigration_documents', applicant_id=applicant_id))

    @app.route('/immigration/documents/<int:document_id>/rename', methods=['POST'])
    def rename_immigration_document(document_id):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        new_name = normalize_whitespace(request.form.get('original_name'))
        applicant_id = request.form.get('applicant_id')
        conn.execute(
            'UPDATE documents SET original_name = ? WHERE id = ? AND model_id = ?',
            (new_name, document_id, model_id),
        )
        maybe_commit(conn)
        return redirect(url_for('immigration_documents', applicant_id=applicant_id))

    @app.route('/immigration/documents/<int:document_id>/delete', methods=['POST'])
    def delete_immigration_document(document_id):
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        applicant_id = request.form.get('applicant_id')
        row = row_to_dict(conn.execute('SELECT * FROM documents WHERE id = ? AND model_id = ?', (document_id, model_id)).fetchone())
        if row:
            absolute_path = resolve_storage_path(row.get('file_path'))
            if os.path.exists(absolute_path):
                try:
                    os.remove(absolute_path)
                except OSError:
                    pass
            conn.execute('DELETE FROM documents WHERE id = ? AND model_id = ?', (document_id, model_id))
            maybe_commit(conn)
        return redirect(url_for('immigration_documents', applicant_id=applicant_id))

    @app.route('/immigration/visa-letters')
    def immigration_visa_letters():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        applicants = [
            row_to_dict(row)
            for row in conn.execute(
                'SELECT * FROM applicants WHERE model_id = ? ORDER BY created_at DESC',
                (model_id,),
            ).fetchall()
        ]
        companies = get_companies(conn, model_id)
        letterheads = get_letterheads(conn, model_id)
        letters = get_recent_visa_letters(conn, model_id, limit=12)
        preview = None
        applicant_id = request.args.get('applicant_id')
        if applicant_id:
            applicant = row_to_dict(conn.execute('SELECT * FROM applicants WHERE id = ? AND model_id = ?', (applicant_id, model_id)).fetchone())
            if applicant:
                company = row_to_dict(conn.execute('SELECT * FROM companies WHERE id = ?', (applicant.get('company_id'),)).fetchone()) or {}
                contact = row_to_dict(conn.execute('SELECT * FROM contacts WHERE id = ?', (applicant.get('contact_id'),)).fetchone()) or {}
                letterhead = row_to_dict(conn.execute('SELECT * FROM letterheads WHERE company_id = ? ORDER BY id DESC LIMIT 1', (applicant.get('company_id'),)).fetchone()) or {}
                preview = build_letter_context(
                    applicant,
                    company,
                    contact,
                    letterhead,
                    applicant.get('reason') or 'Business Meeting',
                    applicant.get('visa_type') or 'Business Visa',
                    bool(letterhead.get('plain_paper')),
                )
        return render_template(
            'immigration/visa_letters.html',
            page_title='Visa Letters',
            applicants=applicants,
            companies=companies,
            letterheads=letterheads,
            reasons=REASONS,
            visa_types=VISA_CATEGORIES,
            letters=letters,
            preview=preview,
        )

    @app.route('/immigration/visa-letters/preview', methods=['POST'])
    def preview_immigration_visa_letter():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        applicant = row_to_dict(conn.execute('SELECT * FROM applicants WHERE id = ? AND model_id = ?', (request.form.get('applicant_id'), model_id)).fetchone()) or {}
        company = row_to_dict(conn.execute('SELECT * FROM companies WHERE id = ? AND model_id = ?', (request.form.get('company_id'), model_id)).fetchone()) or {}
        contact = row_to_dict(conn.execute('SELECT * FROM contacts WHERE id = ?', (applicant.get('contact_id'),)).fetchone()) or {}
        letterhead = row_to_dict(conn.execute('SELECT * FROM letterheads WHERE id = ? AND model_id = ?', (request.form.get('letterhead_id'), model_id)).fetchone()) or {}
        preview = build_letter_context(
            applicant,
            company,
            contact,
            letterhead,
            normalize_whitespace(request.form.get('reason')) or 'Business Meeting',
            normalize_whitespace(request.form.get('visa_type')) or 'Business Visa',
            bool(request.form.get('plain_paper')),
        )
        return jsonify({
            'ok': True,
            'preview': preview,
            'signature_path': build_file_url(letterhead.get('signature_image_path')) if letterhead.get('signature_image_path') else '',
            'background_path': build_file_url(letterhead.get('background_image_path')) if letterhead.get('background_image_path') else '',
        })

    @app.route('/immigration/visa-letters/generate', methods=['POST'])
    def generate_immigration_visa_letter():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        applicant = row_to_dict(conn.execute('SELECT * FROM applicants WHERE id = ? AND model_id = ?', (request.form.get('applicant_id'), model_id)).fetchone()) or {}
        company = row_to_dict(conn.execute('SELECT * FROM companies WHERE id = ? AND model_id = ?', (request.form.get('company_id'), model_id)).fetchone()) or {}
        contact = row_to_dict(conn.execute('SELECT * FROM contacts WHERE id = ?', (applicant.get('contact_id'),)).fetchone()) or {}
        letterhead = row_to_dict(conn.execute('SELECT * FROM letterheads WHERE id = ? AND model_id = ?', (request.form.get('letterhead_id'), model_id)).fetchone()) or {}
        if not applicant or not company:
            return redirect(url_for('immigration_visa_letters'))
        reason = normalize_whitespace(request.form.get('reason')) or 'Business Meeting'
        visa_type = normalize_whitespace(request.form.get('visa_type')) or applicant.get('visa_type') or 'Business Visa'
        plain_paper = 1 if request.form.get('plain_paper') else 0
        preview = build_letter_context(applicant, company, contact, letterhead, reason, visa_type, plain_paper)

        pdf_buffer = io.BytesIO()
        pdf = canvas.Canvas(pdf_buffer, pagesize=A4)
        width, height = A4
        background = letterhead.get('background_image_path')
        if background:
            background_abs = resolve_storage_path(background)
            if os.path.exists(background_abs):
                try:
                    pdf.drawImage(ImageReader(background_abs), 0, 0, width=width, height=height, preserveAspectRatio=False, mask='auto')
                except Exception:
                    pass
        pdf.setFillColorRGB(0.08, 0.12, 0.20)
        pdf.setFont('Helvetica-Bold', 14)
        pdf.drawString(56, height - 88, company.get('name', ''))
        pdf.setFont('Helvetica', 11)
        y = height - 132
        for block in preview['body'].split('\n'):
            if not block:
                y -= 14
                continue
            pdf.drawString(56, y, block[:110])
            y -= 16
        signature_path = letterhead.get('signature_image_path')
        if signature_path:
            signature_abs = resolve_storage_path(signature_path)
            if os.path.exists(signature_abs):
                try:
                    pdf.drawImage(ImageReader(signature_abs), 56, max(y - 24, 90), width=140, height=48, mask='auto')
                except Exception:
                    pass
        pdf.save()
        pdf_buffer.seek(0)

        storage_folder = os.path.join(UPLOAD_ROOT, 'documents')
        os.makedirs(storage_folder, exist_ok=True)
        pdf_name = f"visa_letter_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}.pdf"
        pdf_path = os.path.join(storage_folder, pdf_name)
        with open(pdf_path, 'wb') as handle:
            handle.write(pdf_buffer.getvalue())
        relative_pdf = to_storage_key('documents', pdf_name)

        conn.execute(
            '''
            INSERT INTO visa_letters (
                model_id, applicant_id, company_id, letterhead_id, reason, visa_type,
                nationality, plain_paper, preview_html, pdf_path
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                model_id, applicant.get('id'), company.get('id'), letterhead.get('id'), reason, preview['selected_visa'],
                applicant.get('nationality'), plain_paper, preview['body'], relative_pdf,
            ),
        )
        maybe_commit(conn)

        pdf_buffer.seek(0)
        action = normalize_whitespace(request.form.get('action')) or 'download'
        if action == 'print':
            return send_file(pdf_buffer, mimetype='application/pdf', as_attachment=False, download_name='visa_letter.pdf')
        return send_file(pdf_buffer, mimetype='application/pdf', as_attachment=True, download_name='visa_letter.pdf')

    @app.route('/immigration/automation')
    def immigration_automation():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        current_session = row_to_dict(
            conn.execute(
                'SELECT * FROM automation_sessions WHERE model_id = ? ORDER BY updated_at DESC LIMIT 1',
                (model_id,),
            ).fetchone()
        ) or {'status': 'Disconnected', 'browser_name': 'Edge', 'progress_percent': 0, 'current_page': 'Idle'}
        applicants = [
            row_to_dict(row)
            for row in conn.execute(
                'SELECT id, full_name, reference_number FROM applicants WHERE model_id = ? ORDER BY created_at DESC LIMIT 8',
                (model_id,),
            ).fetchall()
        ]
        return render_template(
            'immigration/automation.html',
            page_title='Automation',
            current_session=current_session,
            applicants=applicants,
        )

    @app.route('/immigration/automation/connect', methods=['POST'])
    def connect_immigration_automation():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        payload = {
            'status': 'Connected',
            'browser_name': 'Microsoft Edge',
            'current_applicant_id': request.form.get('current_applicant_id') or None,
            'current_page': 'Applicant Queue',
            'progress_percent': request.form.get('progress_percent') or 5,
            'details': json.dumps({
                'state': 'Browser connected and ready for portal assistance.',
                'manual_steps': ['CAPTCHA', 'OTP', 'Payment', 'Final submission'],
            }),
        }
        conn.execute(
            '''
            INSERT INTO automation_sessions (
                model_id, status, browser_name, current_applicant_id, current_page, progress_percent, details
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                model_id, payload['status'], payload['browser_name'], payload['current_applicant_id'],
                payload['current_page'], payload['progress_percent'], payload['details'],
            ),
        )
        maybe_commit(conn)
        return redirect(url_for('immigration_automation'))

    @app.route('/immigration/automation/disconnect', methods=['POST'])
    def disconnect_immigration_automation():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        conn.execute(
            '''
            INSERT INTO automation_sessions (
                model_id, status, browser_name, current_applicant_id, current_page, progress_percent, details
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                model_id, 'Disconnected', 'Microsoft Edge', None, 'Idle', 0,
                json.dumps({'state': 'Browser disconnected'}),
            ),
        )
        maybe_commit(conn)
        return redirect(url_for('immigration_automation'))

    @app.route('/immigration/reports')
    def immigration_reports():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        model_id = current_model_id()
        rows = generate_report_rows(conn, model_id, request.args)

        summary = {
            'daily': scalar_value(
                conn.execute("SELECT COUNT(*) AS value FROM applicants WHERE model_id = ? AND DATE(created_at) = DATE(?)", (model_id, datetime.utcnow().strftime('%Y-%m-%d'))).fetchone(),
                0,
            ),
            'weekly': scalar_value(
                conn.execute("SELECT COUNT(*) AS value FROM applicants WHERE model_id = ? AND DATE(created_at) >= DATE(?)", (model_id, (datetime.utcnow() - timedelta(days=7)).strftime('%Y-%m-%d'))).fetchone(),
                0,
            ),
            'monthly': scalar_value(
                conn.execute("SELECT COUNT(*) AS value FROM applicants WHERE model_id = ? AND DATE(created_at) >= DATE(?)", (model_id, (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d'))).fetchone(),
                0,
            ),
        }
        return render_template(
            'immigration/reports.html',
            page_title='Reports',
            rows=rows[:60],
            summary=summary,
            companies=get_companies(conn, model_id),
            visa_types=VISA_CATEGORIES,
            filters=request.args,
        )

    @app.route('/immigration/reports/export')
    def export_immigration_reports():
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        rows = generate_report_rows(conn, current_model_id(), request.args)
        format_type = normalize_whitespace(request.args.get('format')) or 'csv'
        if format_type == 'pdf':
            return send_file(export_rows_to_pdf(rows), mimetype='application/pdf', as_attachment=True, download_name='immigration_report.pdf')
        if format_type == 'xlsx':
            buffer = export_rows_to_xlsx(rows)
            if buffer is not None:
                return send_file(
                    buffer,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name='immigration_report.xlsx',
                )
        return send_file(export_rows_to_csv(rows), mimetype='text/csv', as_attachment=True, download_name='immigration_report.csv')

    @app.route('/immigration/settings', methods=['GET', 'POST'])
    def immigration_settings():
        set_active_workspace(IMMIGRATION_WORKSPACE, persist=False)
        conn = get_db_connection()
        ensure_immigration_ready(conn)
        user_id = session.get('user_id')
        if request.method == 'POST':
            for key in ['travel_default_country', 'travel_default_reason', 'default_company_id', 'default_letterhead_id', 'automation_browser', 'theme_mode']:
                upsert_user_preference(conn, user_id, f'immigration_{key}', request.form.get(key, ''))
            maybe_commit(conn)
            return redirect(url_for('immigration_settings'))

        preferences = {
            'travel_default_country': fetch_user_preference(conn, user_id, 'immigration_travel_default_country', ''),
            'travel_default_reason': fetch_user_preference(conn, user_id, 'immigration_travel_default_reason', ''),
            'default_company_id': fetch_user_preference(conn, user_id, 'immigration_default_company_id', ''),
            'default_letterhead_id': fetch_user_preference(conn, user_id, 'immigration_default_letterhead_id', ''),
            'automation_browser': fetch_user_preference(conn, user_id, 'immigration_automation_browser', 'Microsoft Edge'),
            'theme_mode': fetch_user_preference(conn, user_id, 'immigration_theme_mode', 'system'),
        }
        recent_logs = [
            row_to_dict(row)
            for row in conn.execute(
                '''
                SELECT * FROM audit_logs
                WHERE model_id = ? AND workspace = ?
                ORDER BY created_at DESC
                LIMIT 20
                ''',
                (current_model_id(), IMMIGRATION_WORKSPACE),
            ).fetchall()
        ]
        return render_template(
            'immigration/settings.html',
            page_title='Immigration Settings',
            preferences=preferences,
            companies=get_companies(conn, current_model_id()),
            letterheads=get_letterheads(conn, current_model_id()),
            reasons=REASONS,
            logs=recent_logs,
        )
